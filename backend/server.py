from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File, Response, Form, Request
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.staticfiles import StaticFiles
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
import json
import csv
import io
import base64
import shutil
from twilio.rest import Client as TwilioClient
from emergentintegrations.payments.stripe.checkout import StripeCheckout, CheckoutSessionResponse, CheckoutStatusResponse, CheckoutSessionRequest
import resend
import asyncio
import secrets
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

ROOT_DIR = Path(__file__).parent
UPLOADS_DIR = ROOT_DIR / "uploads"
UPLOADS_DIR.mkdir(exist_ok=True)
(UPLOADS_DIR / "products").mkdir(exist_ok=True)
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Config
JWT_SECRET = os.environ.get('JWT_SECRET', 'cafe-control-secret-key-2024')
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24

# Twilio Config
TWILIO_ACCOUNT_SID = os.environ.get('TWILIO_ACCOUNT_SID')
TWILIO_AUTH_TOKEN = os.environ.get('TWILIO_AUTH_TOKEN')
TWILIO_WHATSAPP_NUMBER = os.environ.get('TWILIO_WHATSAPP_NUMBER')

twilio_client = None
if TWILIO_ACCOUNT_SID and TWILIO_AUTH_TOKEN:
    twilio_client = TwilioClient(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)

# Stripe Config
STRIPE_API_KEY = os.environ.get('STRIPE_API_KEY')

# Resend Config for Password Recovery
RESEND_API_KEY = os.environ.get('RESEND_API_KEY')
SENDER_EMAIL = os.environ.get('SENDER_EMAIL', 'onboarding@resend.dev')
if RESEND_API_KEY:
    resend.api_key = RESEND_API_KEY

# Password Reset Token Expiry (in hours)
PASSWORD_RESET_EXPIRY_HOURS = 1

# Subscription Plans (MXN)
SUBSCRIPTION_PLANS = {
    "plan_1": {"name": "1 Sucursal", "max_branches": 1, "price": 399.00, "currency": "mxn"},
    "plan_2": {"name": "2 Sucursales", "max_branches": 2, "price": 599.00, "currency": "mxn"},
    "plan_3": {"name": "3-5 Sucursales", "max_branches": 5, "price": 799.00, "currency": "mxn"},
    "plan_4": {"name": "5-10 Sucursales", "max_branches": 10, "price": 999.00, "currency": "mxn"},
    "plan_5": {"name": "10-20 Sucursales", "max_branches": 20, "price": 1199.00, "currency": "mxn"},
}

TRIAL_DAYS = 7

app = FastAPI(title="Doré API")
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ============== MODELS ==============

class UserRole:
    ADMIN = "admin"
    GERENTE = "gerente"
    CAJERO = "cajero"
    SUPERADMIN = "superadmin"  # For platform owner

# ============== TENANT/SUBSCRIPTION MODELS ==============

class TenantStatus:
    TRIAL = "trial"
    ACTIVE = "active"
    SUSPENDED = "suspended"
    CANCELLED = "cancelled"

class TenantCreate(BaseModel):
    business_name: str
    owner_name: str
    owner_email: EmailStr
    owner_password: str
    phone: Optional[str] = None

class TenantResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    business_name: str
    owner_email: str
    status: str
    plan_id: Optional[str] = None
    trial_ends_at: Optional[str] = None
    subscription_ends_at: Optional[str] = None
    max_branches: int = 1
    logo_url: Optional[str] = None
    created_at: str

class SubscriptionPlanResponse(BaseModel):
    plan_id: str
    name: str
    max_branches: int
    price: float
    currency: str

class SubscriptionCheckoutRequest(BaseModel):
    plan_id: str
    origin_url: str

class PaymentTransactionResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    tenant_id: str
    plan_id: str
    amount: float
    currency: str
    status: str
    session_id: Optional[str] = None
    created_at: str

# ============== USER MODELS ==============

class UserBase(BaseModel):
    email: EmailStr
    name: str
    role: str = UserRole.CAJERO
    cafeteria_id: Optional[str] = None
    is_active: bool = True
    tenant_id: Optional[str] = None  # Added for multi-tenant

class UserCreate(UserBase):
    password: str

class UserResponse(UserBase):
    model_config = ConfigDict(extra="ignore")
    id: str

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class TokenResponse(BaseModel):
    token: str
    user: UserResponse

# ============== PASSWORD RECOVERY MODELS ==============

class PasswordResetRequest(BaseModel):
    email: EmailStr

class PasswordResetVerify(BaseModel):
    token: str
    new_password: str

class PasswordResetResponse(BaseModel):
    message: str
    success: bool

class CafeteriaBase(BaseModel):
    name: str
    address: str
    phone: Optional[str] = None
    is_active: bool = True

class CafeteriaCreate(CafeteriaBase):
    pass

class CafeteriaResponse(CafeteriaBase):
    model_config = ConfigDict(extra="ignore")
    id: str
    created_at: str

class CategoryBase(BaseModel):
    name: str
    description: Optional[str] = None

class CategoryCreate(CategoryBase):
    pass

class CategoryResponse(CategoryBase):
    model_config = ConfigDict(extra="ignore")
    id: str

# ============== INGREDIENT MODELS ==============

class IngredientBase(BaseModel):
    name: str
    unit: str  # kg, litro, pieza, gramo, ml
    cost_per_unit: float
    supplier_id: Optional[str] = None
    min_stock: float = 10.0
    is_active: bool = True

class IngredientCreate(IngredientBase):
    pass

class IngredientResponse(IngredientBase):
    model_config = ConfigDict(extra="ignore")
    id: str
    supplier_name: Optional[str] = None
    created_at: str

class IngredientInventoryBase(BaseModel):
    ingredient_id: str
    cafeteria_id: str
    quantity: float
    min_stock: float = 10.0

class IngredientInventoryCreate(IngredientInventoryBase):
    pass

class IngredientInventoryResponse(IngredientInventoryBase):
    model_config = ConfigDict(extra="ignore")
    id: str
    ingredient_name: Optional[str] = None
    unit: Optional[str] = None
    cafeteria_name: Optional[str] = None
    is_low_stock: bool = False
    days_until_stockout: Optional[float] = None
    cost_per_unit: float = 0.0

class IngredientMovement(BaseModel):
    inventory_id: str
    quantity: float
    movement_type: str  # "entrada", "salida", "merma", "ajuste", "consumo_venta"
    reason: Optional[str] = None
    sale_id: Optional[str] = None

# ============== RECIPE MODELS ==============

class RecipeIngredient(BaseModel):
    ingredient_id: str
    quantity: float  # cantidad por porción

class RecipeBase(BaseModel):
    product_id: str
    ingredients: List[RecipeIngredient]
    portions: int = 1  # porciones que rinde la receta
    auto_deduct: bool = True  # descuento automático por ventas

class RecipeCreate(RecipeBase):
    pass

class RecipeResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    product_id: str
    product_name: Optional[str] = None
    ingredients: List[dict]
    portions: int
    auto_deduct: bool
    calculated_cost: float = 0.0
    created_at: str

# ============== PRODUCT MODELS (UPDATED) ==============

class ProductBase(BaseModel):
    name: str
    description: Optional[str] = None
    category_id: str
    price: float
    cost: float  # Este será el costo calculado de la receta
    is_active: bool = True
    main_image: Optional[str] = None
    images: List[str] = []  # hasta 3 imágenes adicionales

class ProductCreate(ProductBase):
    pass

class ProductResponse(ProductBase):
    model_config = ConfigDict(extra="ignore")
    id: str
    margin: float = 0.0
    recipe_cost: float = 0.0  # costo calculado de la receta
    has_recipe: bool = False
    created_at: str

class ProductImageUpdate(BaseModel):
    main_image: Optional[str] = None
    images: List[str] = []

# ============== OTHER MODELS ==============

class InventoryItemBase(BaseModel):
    product_id: str
    cafeteria_id: str
    quantity: float
    min_stock: float = 10.0
    unit: str = "unidad"

class InventoryItemCreate(InventoryItemBase):
    pass

class InventoryItemResponse(InventoryItemBase):
    model_config = ConfigDict(extra="ignore")
    id: str
    product_name: Optional[str] = None
    cafeteria_name: Optional[str] = None
    is_low_stock: bool = False

class InventoryMovement(BaseModel):
    inventory_id: str
    quantity: float
    movement_type: str
    reason: Optional[str] = None

class SupplierBase(BaseModel):
    name: str
    contact_name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    is_active: bool = True

class SupplierCreate(SupplierBase):
    pass

class SupplierResponse(SupplierBase):
    model_config = ConfigDict(extra="ignore")
    id: str
    created_at: str

class PurchaseItemBase(BaseModel):
    ingredient_id: Optional[str] = None  # Ahora puede ser ingrediente
    product_id: Optional[str] = None  # o producto terminado
    quantity: float
    unit_cost: float

class PurchaseBase(BaseModel):
    supplier_id: str
    cafeteria_id: str
    items: List[PurchaseItemBase]
    notes: Optional[str] = None

class PurchaseCreate(PurchaseBase):
    pass

class PurchaseResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    supplier_id: str
    supplier_name: Optional[str] = None
    cafeteria_id: str
    cafeteria_name: Optional[str] = None
    items: List[dict]
    total: float
    notes: Optional[str] = None
    created_at: str
    created_by: Optional[str] = None

class SaleItemBase(BaseModel):
    product_id: str
    product_name: str
    quantity: int
    unit_price: float
    subtotal: float

class SaleBase(BaseModel):
    cafeteria_id: str
    items: List[SaleItemBase]
    payment_method: str = "efectivo"
    clip_transaction_id: Optional[str] = None
    notes: Optional[str] = None

class SaleCreate(SaleBase):
    pass

class SaleResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    cafeteria_id: str
    cafeteria_name: Optional[str] = None
    items: List[dict]
    subtotal: float
    tax: float
    total: float
    cost_total: float
    profit: float
    payment_method: str
    clip_transaction_id: Optional[str] = None
    notes: Optional[str] = None
    created_at: str
    created_by: Optional[str] = None

class DashboardStats(BaseModel):
    total_sales_today: float
    total_sales_month: float
    total_profit_today: float
    total_profit_month: float
    sales_count_today: int
    low_stock_alerts: int
    low_ingredient_alerts: int
    top_products: List[dict]
    sales_by_cafeteria: List[dict]
    sales_trend: List[dict]

# ============== AUTH HELPERS ==============

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode(), hashed.encode())

def create_token(user_id: str, email: str, role: str, cafeteria_id: Optional[str] = None, tenant_id: Optional[str] = None) -> str:
    payload = {
        "user_id": user_id,
        "email": email,
        "role": role,
        "cafeteria_id": cafeteria_id,
        "tenant_id": tenant_id,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token inválido")

def require_roles(allowed_roles: List[str]):
    async def role_checker(current_user: dict = Depends(get_current_user)):
        if current_user["role"] not in allowed_roles:
            raise HTTPException(status_code=403, detail="No tienes permisos para esta acción")
        return current_user
    return role_checker

def get_tenant_filter(current_user: dict) -> dict:
    """Get MongoDB filter for tenant isolation"""
    tenant_id = current_user.get("tenant_id")
    if tenant_id:
        return {"tenant_id": tenant_id}
    # For legacy users without tenant, return empty filter (backwards compatibility)
    return {}

async def check_tenant_limit(tenant_id: str, resource_type: str) -> bool:
    """Check if tenant has reached their limit for a resource"""
    if resource_type == "cafeterias":
        tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0})
        if tenant:
            current_count = await db.cafeterias.count_documents({"tenant_id": tenant_id})
            max_branches = tenant.get("max_branches", 1)
            return current_count < max_branches
    return True

# ============== TENANT ROUTES (PUBLIC) ==============

@api_router.get("/plans", response_model=List[SubscriptionPlanResponse])
async def get_plans():
    """Get all subscription plans (public)"""
    return [
        SubscriptionPlanResponse(
            plan_id=plan_id,
            name=plan["name"],
            max_branches=plan["max_branches"],
            price=plan["price"],
            currency=plan["currency"]
        )
        for plan_id, plan in SUBSCRIPTION_PLANS.items()
    ]

@api_router.post("/tenants/register")
async def register_tenant(tenant: TenantCreate):
    """Register a new business/tenant with 7-day trial"""
    # Check if email already exists
    existing = await db.users.find_one({"email": tenant.owner_email})
    if existing:
        raise HTTPException(status_code=400, detail="El email ya está registrado")
    
    existing_tenant = await db.tenants.find_one({"owner_email": tenant.owner_email})
    if existing_tenant:
        raise HTTPException(status_code=400, detail="Ya existe un negocio con este email")
    
    # Create tenant
    tenant_id = str(uuid.uuid4())
    trial_ends = datetime.now(timezone.utc) + timedelta(days=TRIAL_DAYS)
    
    tenant_dict = {
        "id": tenant_id,
        "business_name": tenant.business_name,
        "owner_email": tenant.owner_email,
        "phone": tenant.phone,
        "status": TenantStatus.TRIAL,
        "plan_id": "plan_1",  # Start with basic plan during trial
        "max_branches": 1,
        "logo_url": None,  # Tenant can upload their own logo
        "trial_ends_at": trial_ends.isoformat(),
        "subscription_ends_at": None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.tenants.insert_one(tenant_dict)
    
    # Create owner user as admin
    user_id = str(uuid.uuid4())
    user_dict = {
        "id": user_id,
        "email": tenant.owner_email,
        "name": tenant.owner_name,
        "password": hash_password(tenant.owner_password),
        "role": UserRole.ADMIN,
        "tenant_id": tenant_id,
        "cafeteria_id": None,
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_dict)
    
    # Create token for auto-login
    token = create_token(user_id, tenant.owner_email, UserRole.ADMIN, None, tenant_id)
    
    return {
        "message": "Negocio registrado exitosamente",
        "tenant_id": tenant_id,
        "trial_ends_at": trial_ends.isoformat(),
        "token": token,
        "user": UserResponse(
            id=user_id,
            email=tenant.owner_email,
            name=tenant.owner_name,
            role=UserRole.ADMIN,
            cafeteria_id=None,
            is_active=True,
            tenant_id=tenant_id
        )
    }

@api_router.get("/tenants/me", response_model=TenantResponse)
async def get_my_tenant(current_user: dict = Depends(get_current_user)):
    """Get current user's tenant info"""
    tenant_id = current_user.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=404, detail="No perteneces a ningún negocio")
    
    tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0})
    if not tenant:
        raise HTTPException(status_code=404, detail="Negocio no encontrado")
    
    return TenantResponse(**tenant)

@api_router.post("/tenants/logo")
async def upload_tenant_logo(
    file: UploadFile = File(...),
    current_user: dict = Depends(require_roles([UserRole.ADMIN]))
):
    """Upload tenant logo"""
    tenant_id = current_user.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=400, detail="No perteneces a ningún negocio")
    
    # Validate file type
    allowed_types = ["image/jpeg", "image/png", "image/webp", "image/svg+xml"]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Tipo de archivo no permitido. Use JPEG, PNG, WebP o SVG")
    
    # Validate file size (max 2MB)
    content = await file.read()
    if len(content) > 2 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="El archivo es demasiado grande. Máximo 2MB")
    
    # Create logos directory
    logos_dir = UPLOADS_DIR / "logos"
    logos_dir.mkdir(exist_ok=True)
    
    # Generate filename
    ext = file.filename.split(".")[-1] if "." in file.filename else "png"
    filename = f"logo_{tenant_id}.{ext}"
    file_path = logos_dir / filename
    
    # Delete old logo if exists
    for old_file in logos_dir.glob(f"logo_{tenant_id}.*"):
        old_file.unlink()
    
    # Save file
    with open(file_path, "wb") as f:
        f.write(content)
    
    # Update tenant with logo URL
    logo_url = f"/api/uploads/logos/{filename}"
    await db.tenants.update_one(
        {"id": tenant_id},
        {"$set": {"logo_url": logo_url}}
    )
    
    return {"message": "Logo subido correctamente", "logo_url": logo_url}

@api_router.delete("/tenants/logo")
async def delete_tenant_logo(current_user: dict = Depends(require_roles([UserRole.ADMIN]))):
    """Delete tenant logo"""
    tenant_id = current_user.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=400, detail="No perteneces a ningún negocio")
    
    # Delete logo files
    logos_dir = UPLOADS_DIR / "logos"
    for old_file in logos_dir.glob(f"logo_{tenant_id}.*"):
        old_file.unlink()
    
    # Remove from tenant
    await db.tenants.update_one(
        {"id": tenant_id},
        {"$set": {"logo_url": None}}
    )
    
    return {"message": "Logo eliminado"}

@api_router.get("/uploads/logos/{filename}")
async def get_logo(filename: str):
    """Serve logo files"""
    file_path = UPLOADS_DIR / "logos" / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Logo no encontrado")
    
    ext = filename.split(".")[-1].lower()
    content_types = {
        "jpg": "image/jpeg",
        "jpeg": "image/jpeg",
        "png": "image/png",
        "svg": "image/svg+xml",
        "webp": "image/webp"
    }
    content_type = content_types.get(ext, "application/octet-stream")
    
    with open(file_path, "rb") as f:
        content = f.read()
    
    return Response(content=content, media_type=content_type)

@api_router.put("/tenants/settings")
async def update_tenant_settings(
    business_name: str = Form(None),
    phone: str = Form(None),
    current_user: dict = Depends(require_roles([UserRole.ADMIN]))
):
    """Update tenant settings"""
    tenant_id = current_user.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=400, detail="No perteneces a ningún negocio")
    
    update_data = {}
    if business_name:
        update_data["business_name"] = business_name
    if phone is not None:
        update_data["phone"] = phone
    
    if update_data:
        await db.tenants.update_one({"id": tenant_id}, {"$set": update_data})
    
    tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0})
    return TenantResponse(**tenant)

@api_router.get("/tenants/subscription-status")
async def get_subscription_status(current_user: dict = Depends(get_current_user)):
    """Get subscription status for current tenant"""
    tenant_id = current_user.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=404, detail="No perteneces a ningún negocio")
    
    tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0})
    if not tenant:
        raise HTTPException(status_code=404, detail="Negocio no encontrado")
    
    now = datetime.now(timezone.utc)
    
    # Check trial status
    is_trial = tenant["status"] == TenantStatus.TRIAL
    trial_ends = None
    trial_days_left = 0
    
    if is_trial and tenant.get("trial_ends_at"):
        trial_ends = datetime.fromisoformat(tenant["trial_ends_at"].replace("Z", "+00:00"))
        trial_days_left = max(0, (trial_ends - now).days)
    
    # Check subscription status
    subscription_active = False
    subscription_ends = None
    
    if tenant.get("subscription_ends_at"):
        subscription_ends = datetime.fromisoformat(tenant["subscription_ends_at"].replace("Z", "+00:00"))
        subscription_active = subscription_ends > now
    
    # Count current branches
    branch_count = await db.cafeterias.count_documents({"tenant_id": tenant_id})
    
    plan = SUBSCRIPTION_PLANS.get(tenant.get("plan_id", "plan_1"), SUBSCRIPTION_PLANS["plan_1"])
    
    return {
        "status": tenant["status"],
        "is_trial": is_trial,
        "trial_ends_at": tenant.get("trial_ends_at"),
        "trial_days_left": trial_days_left,
        "subscription_active": subscription_active,
        "subscription_ends_at": tenant.get("subscription_ends_at"),
        "plan_id": tenant.get("plan_id"),
        "plan_name": plan["name"],
        "max_branches": tenant.get("max_branches", 1),
        "current_branches": branch_count,
        "can_add_branch": branch_count < tenant.get("max_branches", 1)
    }

# ============== STRIPE/PAYMENT ROUTES ==============

@api_router.post("/subscription/checkout")
async def create_subscription_checkout(
    request: Request,
    checkout_data: SubscriptionCheckoutRequest,
    current_user: dict = Depends(get_current_user)
):
    """Create Stripe checkout session for subscription"""
    tenant_id = current_user.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=400, detail="No perteneces a ningún negocio")
    
    # Validate plan
    if checkout_data.plan_id not in SUBSCRIPTION_PLANS:
        raise HTTPException(status_code=400, detail="Plan no válido")
    
    plan = SUBSCRIPTION_PLANS[checkout_data.plan_id]
    
    # Initialize Stripe
    host_url = str(request.base_url).rstrip("/")
    webhook_url = f"{host_url}/api/webhook/stripe"
    stripe_checkout = StripeCheckout(api_key=STRIPE_API_KEY, webhook_url=webhook_url)
    
    # Create URLs
    origin = checkout_data.origin_url.rstrip("/")
    success_url = f"{origin}/subscription/success?session_id={{CHECKOUT_SESSION_ID}}"
    cancel_url = f"{origin}/subscription"
    
    # Create checkout session
    checkout_request = CheckoutSessionRequest(
        amount=plan["price"],
        currency=plan["currency"],
        success_url=success_url,
        cancel_url=cancel_url,
        metadata={
            "tenant_id": tenant_id,
            "plan_id": checkout_data.plan_id,
            "user_email": current_user.get("email", "")
        }
    )
    
    session = await stripe_checkout.create_checkout_session(checkout_request)
    
    # Create payment transaction record
    transaction = {
        "id": str(uuid.uuid4()),
        "tenant_id": tenant_id,
        "plan_id": checkout_data.plan_id,
        "amount": plan["price"],
        "currency": plan["currency"],
        "session_id": session.session_id,
        "status": "pending",
        "payment_status": "initiated",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.payment_transactions.insert_one(transaction)
    
    return {
        "checkout_url": session.url,
        "session_id": session.session_id
    }

@api_router.get("/subscription/checkout/status/{session_id}")
async def get_checkout_status(
    request: Request,
    session_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Check payment status and activate subscription if paid"""
    tenant_id = current_user.get("tenant_id")
    
    # Find transaction
    transaction = await db.payment_transactions.find_one({"session_id": session_id}, {"_id": 0})
    if not transaction:
        raise HTTPException(status_code=404, detail="Transacción no encontrada")
    
    if transaction["tenant_id"] != tenant_id:
        raise HTTPException(status_code=403, detail="No autorizado")
    
    # Already processed?
    if transaction.get("payment_status") == "paid":
        return {
            "status": "complete",
            "payment_status": "paid",
            "message": "Suscripción ya activada"
        }
    
    # Check with Stripe
    host_url = str(request.base_url).rstrip("/")
    webhook_url = f"{host_url}/api/webhook/stripe"
    stripe_checkout = StripeCheckout(api_key=STRIPE_API_KEY, webhook_url=webhook_url)
    
    status = await stripe_checkout.get_checkout_status(session_id)
    
    # Update transaction
    await db.payment_transactions.update_one(
        {"session_id": session_id},
        {"$set": {"status": status.status, "payment_status": status.payment_status}}
    )
    
    # If paid, activate subscription
    if status.payment_status == "paid":
        plan_id = transaction["plan_id"]
        plan = SUBSCRIPTION_PLANS.get(plan_id, SUBSCRIPTION_PLANS["plan_1"])
        
        # Set subscription for 30 days
        subscription_ends = datetime.now(timezone.utc) + timedelta(days=30)
        
        await db.tenants.update_one(
            {"id": tenant_id},
            {"$set": {
                "status": TenantStatus.ACTIVE,
                "plan_id": plan_id,
                "max_branches": plan["max_branches"],
                "subscription_ends_at": subscription_ends.isoformat()
            }}
        )
        
        return {
            "status": status.status,
            "payment_status": status.payment_status,
            "message": "¡Suscripción activada exitosamente!",
            "subscription_ends_at": subscription_ends.isoformat()
        }
    
    return {
        "status": status.status,
        "payment_status": status.payment_status,
        "message": "Pago pendiente"
    }

@api_router.post("/webhook/stripe")
async def stripe_webhook(request: Request):
    """Handle Stripe webhooks"""
    try:
        body = await request.body()
        signature = request.headers.get("Stripe-Signature")
        
        host_url = str(request.base_url).rstrip("/")
        webhook_url = f"{host_url}/api/webhook/stripe"
        stripe_checkout = StripeCheckout(api_key=STRIPE_API_KEY, webhook_url=webhook_url)
        
        webhook_response = await stripe_checkout.handle_webhook(body, signature)
        
        if webhook_response.payment_status == "paid":
            session_id = webhook_response.session_id
            metadata = webhook_response.metadata
            
            # Update transaction
            await db.payment_transactions.update_one(
                {"session_id": session_id},
                {"$set": {"status": "complete", "payment_status": "paid"}}
            )
            
            # Activate subscription
            tenant_id = metadata.get("tenant_id")
            plan_id = metadata.get("plan_id", "plan_1")
            
            if tenant_id:
                plan = SUBSCRIPTION_PLANS.get(plan_id, SUBSCRIPTION_PLANS["plan_1"])
                subscription_ends = datetime.now(timezone.utc) + timedelta(days=30)
                
                await db.tenants.update_one(
                    {"id": tenant_id},
                    {"$set": {
                        "status": TenantStatus.ACTIVE,
                        "plan_id": plan_id,
                        "max_branches": plan["max_branches"],
                        "subscription_ends_at": subscription_ends.isoformat()
                    }}
                )
        
        return {"status": "ok"}
    except Exception as e:
        logger.error(f"Webhook error: {str(e)}")
        return {"status": "error", "message": str(e)}

# ============== SUPER ADMIN ROUTES ==============

@api_router.get("/admin/tenants")
async def get_all_tenants(current_user: dict = Depends(get_current_user)):
    """Get all tenants (super admin only)"""
    if current_user.get("role") != UserRole.SUPERADMIN:
        raise HTTPException(status_code=403, detail="Solo super admin")
    
    tenants = await db.tenants.find({}, {"_id": 0}).to_list(10000)
    
    # Add branch count for each tenant
    for tenant in tenants:
        tenant["branch_count"] = await db.cafeterias.count_documents({"tenant_id": tenant["id"]})
        tenant["user_count"] = await db.users.count_documents({"tenant_id": tenant["id"]})
    
    return tenants

@api_router.get("/admin/stats")
async def get_admin_stats(current_user: dict = Depends(get_current_user)):
    """Get platform stats (super admin only)"""
    if current_user.get("role") != UserRole.SUPERADMIN:
        raise HTTPException(status_code=403, detail="Solo super admin")
    
    total_tenants = await db.tenants.count_documents({})
    active_tenants = await db.tenants.count_documents({"status": TenantStatus.ACTIVE})
    trial_tenants = await db.tenants.count_documents({"status": TenantStatus.TRIAL})
    total_revenue = 0
    
    # Calculate revenue
    paid_transactions = await db.payment_transactions.find({"payment_status": "paid"}, {"_id": 0}).to_list(10000)
    total_revenue = sum(t.get("amount", 0) for t in paid_transactions)
    
    return {
        "total_tenants": total_tenants,
        "active_tenants": active_tenants,
        "trial_tenants": trial_tenants,
        "total_revenue": total_revenue,
        "currency": "MXN"
    }

# ============== AUTH ROUTES ==============

@api_router.post("/auth/register", response_model=TokenResponse)
async def register(user: UserCreate):
    existing = await db.users.find_one({"email": user.email})
    if existing:
        raise HTTPException(status_code=400, detail="El email ya está registrado")
    
    user_id = str(uuid.uuid4())
    user_dict = user.model_dump()
    user_dict["id"] = user_id
    user_dict["password"] = hash_password(user.password)
    user_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.users.insert_one(user_dict)
    
    token = create_token(user_id, user.email, user.role, user.cafeteria_id)
    return TokenResponse(
        token=token,
        user=UserResponse(id=user_id, email=user.email, name=user.name, role=user.role, cafeteria_id=user.cafeteria_id)
    )

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(credentials: UserLogin):
    user = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    if not user or not verify_password(credentials.password, user["password"]):
        raise HTTPException(status_code=401, detail="Credenciales inválidas")
    
    if not user.get("is_active", True):
        raise HTTPException(status_code=401, detail="Usuario desactivado")
    
    # Check tenant status if user belongs to a tenant
    tenant_id = user.get("tenant_id")
    if tenant_id:
        tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0})
        if tenant:
            now = datetime.now(timezone.utc)
            # Check if trial expired
            if tenant["status"] == TenantStatus.TRIAL:
                trial_ends = datetime.fromisoformat(tenant["trial_ends_at"].replace("Z", "+00:00"))
                if trial_ends < now:
                    # Update tenant status
                    await db.tenants.update_one({"id": tenant_id}, {"$set": {"status": TenantStatus.SUSPENDED}})
            # Check if subscription expired
            elif tenant["status"] == TenantStatus.ACTIVE and tenant.get("subscription_ends_at"):
                sub_ends = datetime.fromisoformat(tenant["subscription_ends_at"].replace("Z", "+00:00"))
                if sub_ends < now:
                    await db.tenants.update_one({"id": tenant_id}, {"$set": {"status": TenantStatus.SUSPENDED}})
    
    token = create_token(user["id"], user["email"], user["role"], user.get("cafeteria_id"), tenant_id)
    return TokenResponse(
        token=token,
        user=UserResponse(
            id=user["id"],
            email=user["email"],
            name=user["name"],
            role=user["role"],
            cafeteria_id=user.get("cafeteria_id"),
            is_active=user.get("is_active", True),
            tenant_id=tenant_id
        )
    )

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(current_user: dict = Depends(get_current_user)):
    user = await db.users.find_one({"id": current_user["user_id"]}, {"_id": 0, "password": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    return UserResponse(**user)

# ============== USER ROUTES ==============

@api_router.get("/users", response_model=List[UserResponse])
async def get_users(current_user: dict = Depends(require_roles([UserRole.ADMIN]))):
    tenant_filter = get_tenant_filter(current_user)
    users = await db.users.find(tenant_filter, {"_id": 0, "password": 0}).to_list(1000)
    return [UserResponse(**u) for u in users]

@api_router.post("/users", response_model=UserResponse)
async def create_user(user: UserCreate, current_user: dict = Depends(require_roles([UserRole.ADMIN]))):
    existing = await db.users.find_one({"email": user.email})
    if existing:
        raise HTTPException(status_code=400, detail="El email ya está registrado")
    
    user_id = str(uuid.uuid4())
    user_dict = user.model_dump()
    user_dict["id"] = user_id
    user_dict["password"] = hash_password(user.password)
    user_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    user_dict["tenant_id"] = current_user.get("tenant_id")  # Assign to same tenant
    
    await db.users.insert_one(user_dict)
    return UserResponse(id=user_id, email=user.email, name=user.name, role=user.role, cafeteria_id=user.cafeteria_id, tenant_id=user_dict["tenant_id"])

@api_router.put("/users/{user_id}", response_model=UserResponse)
async def update_user(user_id: str, user: UserBase, current_user: dict = Depends(require_roles([UserRole.ADMIN]))):
    tenant_filter = get_tenant_filter(current_user)
    tenant_filter["id"] = user_id
    result = await db.users.update_one(tenant_filter, {"$set": user.model_dump()})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    updated = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
    return UserResponse(**updated)

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(require_roles([UserRole.ADMIN]))):
    tenant_filter = get_tenant_filter(current_user)
    tenant_filter["id"] = user_id
    result = await db.users.delete_one(tenant_filter)
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    return {"message": "Usuario eliminado"}

# ============== CAFETERIA ROUTES ==============

@api_router.get("/cafeterias", response_model=List[CafeteriaResponse])
async def get_cafeterias(current_user: dict = Depends(get_current_user)):
    tenant_filter = get_tenant_filter(current_user)
    cafeterias = await db.cafeterias.find(tenant_filter, {"_id": 0}).to_list(100)
    return [CafeteriaResponse(**c) for c in cafeterias]

@api_router.post("/cafeterias", response_model=CafeteriaResponse)
async def create_cafeteria(cafeteria: CafeteriaCreate, current_user: dict = Depends(require_roles([UserRole.ADMIN]))):
    tenant_id = current_user.get("tenant_id")
    
    # Check tenant limit
    if tenant_id and not await check_tenant_limit(tenant_id, "cafeterias"):
        raise HTTPException(status_code=403, detail="Has alcanzado el límite de sucursales de tu plan. Actualiza tu suscripción para agregar más.")
    
    cafeteria_id = str(uuid.uuid4())
    cafeteria_dict = cafeteria.model_dump()
    cafeteria_dict["id"] = cafeteria_id
    cafeteria_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    cafeteria_dict["tenant_id"] = tenant_id
    
    await db.cafeterias.insert_one(cafeteria_dict)
    return CafeteriaResponse(**cafeteria_dict)

@api_router.put("/cafeterias/{cafeteria_id}", response_model=CafeteriaResponse)
async def update_cafeteria(cafeteria_id: str, cafeteria: CafeteriaBase, current_user: dict = Depends(require_roles([UserRole.ADMIN]))):
    tenant_filter = get_tenant_filter(current_user)
    tenant_filter["id"] = cafeteria_id
    result = await db.cafeterias.update_one(tenant_filter, {"$set": cafeteria.model_dump()})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Cafetería no encontrada")
    updated = await db.cafeterias.find_one({"id": cafeteria_id}, {"_id": 0})
    return CafeteriaResponse(**updated)

@api_router.delete("/cafeterias/{cafeteria_id}")
async def delete_cafeteria(cafeteria_id: str, current_user: dict = Depends(require_roles([UserRole.ADMIN]))):
    tenant_filter = get_tenant_filter(current_user)
    tenant_filter["id"] = cafeteria_id
    result = await db.cafeterias.delete_one(tenant_filter)
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Cafetería no encontrada")
    return {"message": "Cafetería eliminada"}

# ============== CATEGORY ROUTES ==============

@api_router.get("/categories", response_model=List[CategoryResponse])
async def get_categories(current_user: dict = Depends(get_current_user)):
    tenant_filter = get_tenant_filter(current_user)
    categories = await db.categories.find(tenant_filter, {"_id": 0}).to_list(100)
    return [CategoryResponse(**c) for c in categories]

@api_router.post("/categories", response_model=CategoryResponse)
async def create_category(category: CategoryCreate, current_user: dict = Depends(require_roles([UserRole.ADMIN, UserRole.GERENTE]))):
    category_id = str(uuid.uuid4())
    category_dict = category.model_dump()
    category_dict["id"] = category_id
    category_dict["tenant_id"] = current_user.get("tenant_id")
    
    await db.categories.insert_one(category_dict)
    return CategoryResponse(**category_dict)

@api_router.delete("/categories/{category_id}")
async def delete_category(category_id: str, current_user: dict = Depends(require_roles([UserRole.ADMIN]))):
    tenant_filter = get_tenant_filter(current_user)
    tenant_filter["id"] = category_id
    result = await db.categories.delete_one(tenant_filter)
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Categoría no encontrada")
    return {"message": "Categoría eliminada"}

# ============== INGREDIENT ROUTES ==============

@api_router.get("/ingredients", response_model=List[IngredientResponse])
async def get_ingredients(current_user: dict = Depends(get_current_user)):
    tenant_filter = get_tenant_filter(current_user)
    ingredients = await db.ingredients.find(tenant_filter, {"_id": 0}).to_list(1000)
    suppliers = {s["id"]: s["name"] for s in await db.suppliers.find(tenant_filter, {"_id": 0, "id": 1, "name": 1}).to_list(100)}
    
    result = []
    for ing in ingredients:
        ing["supplier_name"] = suppliers.get(ing.get("supplier_id"), None)
        result.append(IngredientResponse(**ing))
    return result

@api_router.post("/ingredients", response_model=IngredientResponse)
async def create_ingredient(ingredient: IngredientCreate, current_user: dict = Depends(require_roles([UserRole.ADMIN, UserRole.GERENTE]))):
    ingredient_id = str(uuid.uuid4())
    ingredient_dict = ingredient.model_dump()
    ingredient_dict["id"] = ingredient_id
    ingredient_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    ingredient_dict["tenant_id"] = current_user.get("tenant_id")
    
    await db.ingredients.insert_one(ingredient_dict)
    
    supplier_name = None
    if ingredient.supplier_id:
        supplier = await db.suppliers.find_one({"id": ingredient.supplier_id}, {"_id": 0, "name": 1})
        supplier_name = supplier["name"] if supplier else None
    
    return IngredientResponse(**ingredient_dict, supplier_name=supplier_name)

@api_router.put("/ingredients/{ingredient_id}", response_model=IngredientResponse)
async def update_ingredient(ingredient_id: str, ingredient: IngredientBase, current_user: dict = Depends(require_roles([UserRole.ADMIN, UserRole.GERENTE]))):
    tenant_filter = get_tenant_filter(current_user)
    tenant_filter["id"] = ingredient_id
    result = await db.ingredients.update_one(tenant_filter, {"$set": ingredient.model_dump()})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Ingrediente no encontrado")
    
    updated = await db.ingredients.find_one({"id": ingredient_id}, {"_id": 0})
    supplier_name = None
    if updated.get("supplier_id"):
        supplier = await db.suppliers.find_one({"id": updated["supplier_id"]}, {"_id": 0, "name": 1})
        supplier_name = supplier["name"] if supplier else None
    
    return IngredientResponse(**updated, supplier_name=supplier_name)

@api_router.delete("/ingredients/{ingredient_id}")
async def delete_ingredient(ingredient_id: str, current_user: dict = Depends(require_roles([UserRole.ADMIN]))):
    tenant_filter = get_tenant_filter(current_user)
    tenant_filter["id"] = ingredient_id
    result = await db.ingredients.delete_one(tenant_filter)
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Ingrediente no encontrado")
    return {"message": "Ingrediente eliminado"}

# ============== INGREDIENT INVENTORY ROUTES ==============

@api_router.get("/ingredient-inventory", response_model=List[IngredientInventoryResponse])
async def get_ingredient_inventory(cafeteria_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    tenant_filter = get_tenant_filter(current_user)
    query = {**tenant_filter}
    if cafeteria_id:
        query["cafeteria_id"] = cafeteria_id
    elif current_user["role"] == UserRole.GERENTE and current_user.get("cafeteria_id"):
        query["cafeteria_id"] = current_user["cafeteria_id"]
    
    inventory = await db.ingredient_inventory.find(query, {"_id": 0}).to_list(1000)
    
    ingredients = {i["id"]: i for i in await db.ingredients.find(tenant_filter, {"_id": 0}).to_list(1000)}
    cafeterias = {c["id"]: c["name"] for c in await db.cafeterias.find(tenant_filter, {"_id": 0, "id": 1, "name": 1}).to_list(100)}
    
    # Calculate average daily consumption from last 30 days
    thirty_days_ago = (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()
    mov_query = {**tenant_filter, "movement_type": "consumo_venta", "created_at": {"$gte": thirty_days_ago}}
    movements = await db.ingredient_movements.find(mov_query, {"_id": 0}).to_list(10000)
    
    daily_consumption = {}
    for mov in movements:
        key = f"{mov['ingredient_id']}_{mov.get('cafeteria_id', '')}"
        if key not in daily_consumption:
            daily_consumption[key] = 0
        daily_consumption[key] += mov["quantity"]
    
    result = []
    for item in inventory:
        ing = ingredients.get(item["ingredient_id"], {})
        item["ingredient_name"] = ing.get("name", "Desconocido")
        item["unit"] = ing.get("unit", "unidad")
        item["cost_per_unit"] = ing.get("cost_per_unit", 0)
        item["cafeteria_name"] = cafeterias.get(item["cafeteria_id"], "Desconocida")
        item["is_low_stock"] = item["quantity"] <= item["min_stock"]
        
        # Calculate days until stockout
        key = f"{item['ingredient_id']}_{item['cafeteria_id']}"
        avg_daily = daily_consumption.get(key, 0) / 30 if daily_consumption.get(key) else 0
        item["days_until_stockout"] = round(item["quantity"] / avg_daily, 1) if avg_daily > 0 else None
        
        result.append(IngredientInventoryResponse(**item))
    
    return result

@api_router.post("/ingredient-inventory", response_model=IngredientInventoryResponse)
async def create_ingredient_inventory(item: IngredientInventoryCreate, current_user: dict = Depends(require_roles([UserRole.ADMIN, UserRole.GERENTE]))):
    existing = await db.ingredient_inventory.find_one({
        "ingredient_id": item.ingredient_id,
        "cafeteria_id": item.cafeteria_id
    })
    if existing:
        raise HTTPException(status_code=400, detail="Ya existe inventario para este ingrediente en esta cafetería")
    
    item_id = str(uuid.uuid4())
    item_dict = item.model_dump()
    item_dict["id"] = item_id
    item_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.ingredient_inventory.insert_one(item_dict)
    
    ingredient = await db.ingredients.find_one({"id": item.ingredient_id}, {"_id": 0})
    cafeteria = await db.cafeterias.find_one({"id": item.cafeteria_id}, {"_id": 0, "name": 1})
    
    return IngredientInventoryResponse(
        **item_dict,
        ingredient_name=ingredient["name"] if ingredient else "Desconocido",
        unit=ingredient.get("unit", "unidad") if ingredient else "unidad",
        cost_per_unit=ingredient.get("cost_per_unit", 0) if ingredient else 0,
        cafeteria_name=cafeteria["name"] if cafeteria else "Desconocida",
        is_low_stock=item.quantity <= item.min_stock
    )

@api_router.post("/ingredient-inventory/movement")
async def record_ingredient_movement(movement: IngredientMovement, current_user: dict = Depends(require_roles([UserRole.ADMIN, UserRole.GERENTE]))):
    inventory_item = await db.ingredient_inventory.find_one({"id": movement.inventory_id})
    if not inventory_item:
        raise HTTPException(status_code=404, detail="Item de inventario no encontrado")
    
    new_quantity = inventory_item["quantity"]
    if movement.movement_type in ["entrada"]:
        new_quantity += movement.quantity
    elif movement.movement_type in ["salida", "merma", "consumo_venta"]:
        new_quantity -= movement.quantity
    else:  # ajuste
        new_quantity = movement.quantity
    
    if new_quantity < 0:
        raise HTTPException(status_code=400, detail="No hay suficiente stock")
    
    await db.ingredient_inventory.update_one({"id": movement.inventory_id}, {"$set": {"quantity": new_quantity}})
    
    movement_log = {
        "id": str(uuid.uuid4()),
        "inventory_id": movement.inventory_id,
        "ingredient_id": inventory_item["ingredient_id"],
        "cafeteria_id": inventory_item["cafeteria_id"],
        "quantity": movement.quantity,
        "movement_type": movement.movement_type,
        "reason": movement.reason,
        "sale_id": movement.sale_id,
        "previous_quantity": inventory_item["quantity"],
        "new_quantity": new_quantity,
        "created_by": current_user["user_id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.ingredient_movements.insert_one(movement_log)
    
    return {"message": "Movimiento registrado", "new_quantity": new_quantity}

@api_router.get("/ingredient-inventory/alerts")
async def get_ingredient_alerts(cafeteria_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """Get ingredients that need restocking based on consumption rate"""
    query = {}
    if cafeteria_id:
        query["cafeteria_id"] = cafeteria_id
    
    inventory = await db.ingredient_inventory.find(query, {"_id": 0}).to_list(1000)
    ingredients = {i["id"]: i for i in await db.ingredients.find({}, {"_id": 0}).to_list(1000)}
    cafeterias = {c["id"]: c["name"] for c in await db.cafeterias.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(100)}
    
    # Calculate consumption
    thirty_days_ago = (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()
    movements = await db.ingredient_movements.find({
        "movement_type": "consumo_venta",
        "created_at": {"$gte": thirty_days_ago}
    }, {"_id": 0}).to_list(10000)
    
    daily_consumption = {}
    for mov in movements:
        key = f"{mov['ingredient_id']}_{mov.get('cafeteria_id', '')}"
        if key not in daily_consumption:
            daily_consumption[key] = 0
        daily_consumption[key] += mov["quantity"]
    
    alerts = []
    for item in inventory:
        key = f"{item['ingredient_id']}_{item['cafeteria_id']}"
        avg_daily = daily_consumption.get(key, 0) / 30 if daily_consumption.get(key) else 0
        days_left = item["quantity"] / avg_daily if avg_daily > 0 else float('inf')
        
        ing = ingredients.get(item["ingredient_id"], {})
        
        # Alert if less than 7 days of stock or below minimum
        if days_left < 7 or item["quantity"] <= item["min_stock"]:
            alerts.append({
                "ingredient_id": item["ingredient_id"],
                "ingredient_name": ing.get("name", "Desconocido"),
                "unit": ing.get("unit", "unidad"),
                "cafeteria_id": item["cafeteria_id"],
                "cafeteria_name": cafeterias.get(item["cafeteria_id"], "Desconocida"),
                "current_stock": item["quantity"],
                "min_stock": item["min_stock"],
                "avg_daily_consumption": round(avg_daily, 2),
                "days_until_stockout": round(days_left, 1) if days_left != float('inf') else None,
                "suggested_order": round(avg_daily * 14 - item["quantity"], 2) if avg_daily > 0 else item["min_stock"] * 2,
                "alert_type": "critical" if days_left < 3 else "warning"
            })
    
    return sorted(alerts, key=lambda x: x.get("days_until_stockout") or float('inf'))

# ============== RECIPE ROUTES ==============

@api_router.get("/recipes", response_model=List[RecipeResponse])
async def get_recipes(current_user: dict = Depends(get_current_user)):
    recipes = await db.recipes.find({}, {"_id": 0}).to_list(1000)
    products = {p["id"]: p["name"] for p in await db.products.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(1000)}
    ingredients = {i["id"]: i for i in await db.ingredients.find({}, {"_id": 0}).to_list(1000)}
    
    result = []
    for recipe in recipes:
        recipe["product_name"] = products.get(recipe["product_id"], "Desconocido")
        
        # Enrich ingredients and calculate cost
        calculated_cost = 0
        enriched_ingredients = []
        for ing_item in recipe.get("ingredients", []):
            ing = ingredients.get(ing_item["ingredient_id"], {})
            cost = ing.get("cost_per_unit", 0) * ing_item["quantity"]
            calculated_cost += cost
            enriched_ingredients.append({
                "ingredient_id": ing_item["ingredient_id"],
                "ingredient_name": ing.get("name", "Desconocido"),
                "unit": ing.get("unit", "unidad"),
                "quantity": ing_item["quantity"],
                "cost_per_unit": ing.get("cost_per_unit", 0),
                "subtotal": round(cost, 2)
            })
        
        recipe["ingredients"] = enriched_ingredients
        recipe["calculated_cost"] = round(calculated_cost / recipe.get("portions", 1), 2)
        result.append(RecipeResponse(**recipe))
    
    return result

@api_router.get("/recipes/product/{product_id}", response_model=Optional[RecipeResponse])
async def get_recipe_by_product(product_id: str, current_user: dict = Depends(get_current_user)):
    recipe = await db.recipes.find_one({"product_id": product_id}, {"_id": 0})
    if not recipe:
        return None
    
    product = await db.products.find_one({"id": product_id}, {"_id": 0, "name": 1})
    ingredients = {i["id"]: i for i in await db.ingredients.find({}, {"_id": 0}).to_list(1000)}
    
    recipe["product_name"] = product["name"] if product else "Desconocido"
    
    calculated_cost = 0
    enriched_ingredients = []
    for ing_item in recipe.get("ingredients", []):
        ing = ingredients.get(ing_item["ingredient_id"], {})
        cost = ing.get("cost_per_unit", 0) * ing_item["quantity"]
        calculated_cost += cost
        enriched_ingredients.append({
            "ingredient_id": ing_item["ingredient_id"],
            "ingredient_name": ing.get("name", "Desconocido"),
            "unit": ing.get("unit", "unidad"),
            "quantity": ing_item["quantity"],
            "cost_per_unit": ing.get("cost_per_unit", 0),
            "subtotal": round(cost, 2)
        })
    
    recipe["ingredients"] = enriched_ingredients
    recipe["calculated_cost"] = round(calculated_cost / recipe.get("portions", 1), 2)
    
    return RecipeResponse(**recipe)

@api_router.post("/recipes", response_model=RecipeResponse)
async def create_recipe(recipe: RecipeCreate, current_user: dict = Depends(require_roles([UserRole.ADMIN, UserRole.GERENTE]))):
    # Check if recipe already exists for this product
    existing = await db.recipes.find_one({"product_id": recipe.product_id})
    if existing:
        raise HTTPException(status_code=400, detail="Ya existe una receta para este producto. Use PUT para actualizar.")
    
    recipe_id = str(uuid.uuid4())
    recipe_dict = recipe.model_dump()
    recipe_dict["id"] = recipe_id
    recipe_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.recipes.insert_one(recipe_dict)
    
    # Calculate cost and update product
    ingredients = {i["id"]: i for i in await db.ingredients.find({}, {"_id": 0}).to_list(1000)}
    calculated_cost = 0
    enriched_ingredients = []
    
    for ing_item in recipe.ingredients:
        ing = ingredients.get(ing_item.ingredient_id, {})
        cost = ing.get("cost_per_unit", 0) * ing_item.quantity
        calculated_cost += cost
        enriched_ingredients.append({
            "ingredient_id": ing_item.ingredient_id,
            "ingredient_name": ing.get("name", "Desconocido"),
            "unit": ing.get("unit", "unidad"),
            "quantity": ing_item.quantity,
            "cost_per_unit": ing.get("cost_per_unit", 0),
            "subtotal": round(cost, 2)
        })
    
    cost_per_portion = calculated_cost / recipe.portions
    
    # Update product cost
    await db.products.update_one({"id": recipe.product_id}, {"$set": {"cost": round(cost_per_portion, 2)}})
    
    product = await db.products.find_one({"id": recipe.product_id}, {"_id": 0, "name": 1})
    
    return RecipeResponse(
        id=recipe_id,
        product_id=recipe.product_id,
        product_name=product["name"] if product else "Desconocido",
        ingredients=enriched_ingredients,
        portions=recipe.portions,
        auto_deduct=recipe.auto_deduct,
        calculated_cost=round(cost_per_portion, 2),
        created_at=recipe_dict["created_at"]
    )

@api_router.put("/recipes/{recipe_id}", response_model=RecipeResponse)
async def update_recipe(recipe_id: str, recipe: RecipeCreate, current_user: dict = Depends(require_roles([UserRole.ADMIN, UserRole.GERENTE]))):
    existing = await db.recipes.find_one({"id": recipe_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Receta no encontrada")
    
    recipe_dict = recipe.model_dump()
    recipe_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.recipes.update_one({"id": recipe_id}, {"$set": recipe_dict})
    
    # Recalculate cost
    ingredients = {i["id"]: i for i in await db.ingredients.find({}, {"_id": 0}).to_list(1000)}
    calculated_cost = 0
    enriched_ingredients = []
    
    for ing_item in recipe.ingredients:
        ing = ingredients.get(ing_item.ingredient_id, {})
        cost = ing.get("cost_per_unit", 0) * ing_item.quantity
        calculated_cost += cost
        enriched_ingredients.append({
            "ingredient_id": ing_item.ingredient_id,
            "ingredient_name": ing.get("name", "Desconocido"),
            "unit": ing.get("unit", "unidad"),
            "quantity": ing_item.quantity,
            "cost_per_unit": ing.get("cost_per_unit", 0),
            "subtotal": round(cost, 2)
        })
    
    cost_per_portion = calculated_cost / recipe.portions
    await db.products.update_one({"id": recipe.product_id}, {"$set": {"cost": round(cost_per_portion, 2)}})
    
    product = await db.products.find_one({"id": recipe.product_id}, {"_id": 0, "name": 1})
    
    return RecipeResponse(
        id=recipe_id,
        product_id=recipe.product_id,
        product_name=product["name"] if product else "Desconocido",
        ingredients=enriched_ingredients,
        portions=recipe.portions,
        auto_deduct=recipe.auto_deduct,
        calculated_cost=round(cost_per_portion, 2),
        created_at=existing.get("created_at", "")
    )

@api_router.delete("/recipes/{recipe_id}")
async def delete_recipe(recipe_id: str, current_user: dict = Depends(require_roles([UserRole.ADMIN]))):
    result = await db.recipes.delete_one({"id": recipe_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Receta no encontrada")
    return {"message": "Receta eliminada"}

# ============== PRODUCT ROUTES (UPDATED) ==============

@api_router.get("/products", response_model=List[ProductResponse])
async def get_products(current_user: dict = Depends(get_current_user)):
    tenant_filter = get_tenant_filter(current_user)
    products = await db.products.find(tenant_filter, {"_id": 0}).to_list(1000)
    recipes = {r["product_id"]: r for r in await db.recipes.find(tenant_filter, {"_id": 0}).to_list(1000)}
    ingredients = {i["id"]: i for i in await db.ingredients.find(tenant_filter, {"_id": 0}).to_list(1000)}
    
    result = []
    for p in products:
        recipe = recipes.get(p["id"])
        recipe_cost = 0
        has_recipe = False
        
        if recipe:
            has_recipe = True
            for ing_item in recipe.get("ingredients", []):
                ing = ingredients.get(ing_item["ingredient_id"], {})
                recipe_cost += ing.get("cost_per_unit", 0) * ing_item["quantity"]
            recipe_cost = recipe_cost / recipe.get("portions", 1)
        
        margin = ((p["price"] - p["cost"]) / p["price"] * 100) if p["price"] > 0 else 0
        result.append(ProductResponse(
            **p,
            margin=round(margin, 2),
            recipe_cost=round(recipe_cost, 2),
            has_recipe=has_recipe
        ))
    return result

@api_router.post("/products", response_model=ProductResponse)
async def create_product(product: ProductCreate, current_user: dict = Depends(require_roles([UserRole.ADMIN, UserRole.GERENTE]))):
    product_id = str(uuid.uuid4())
    product_dict = product.model_dump()
    product_dict["id"] = product_id
    product_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    product_dict["tenant_id"] = current_user.get("tenant_id")
    
    await db.products.insert_one(product_dict)
    margin = ((product.price - product.cost) / product.price * 100) if product.price > 0 else 0
    return ProductResponse(**product_dict, margin=round(margin, 2), recipe_cost=0, has_recipe=False)

@api_router.put("/products/{product_id}", response_model=ProductResponse)
async def update_product(product_id: str, product: ProductBase, current_user: dict = Depends(require_roles([UserRole.ADMIN, UserRole.GERENTE]))):
    tenant_filter = get_tenant_filter(current_user)
    tenant_filter["id"] = product_id
    result = await db.products.update_one(tenant_filter, {"$set": product.model_dump()})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    updated = await db.products.find_one({"id": product_id}, {"_id": 0})
    
    recipe = await db.recipes.find_one({"product_id": product_id}, {"_id": 0})
    recipe_cost = 0
    has_recipe = False
    
    if recipe:
        has_recipe = True
        ingredients = {i["id"]: i for i in await db.ingredients.find(get_tenant_filter(current_user), {"_id": 0}).to_list(1000)}
        for ing_item in recipe.get("ingredients", []):
            ing = ingredients.get(ing_item["ingredient_id"], {})
            recipe_cost += ing.get("cost_per_unit", 0) * ing_item["quantity"]
        recipe_cost = recipe_cost / recipe.get("portions", 1)
    
    margin = ((updated["price"] - updated["cost"]) / updated["price"] * 100) if updated["price"] > 0 else 0
    return ProductResponse(**updated, margin=round(margin, 2), recipe_cost=round(recipe_cost, 2), has_recipe=has_recipe)

@api_router.put("/products/{product_id}/images")
async def update_product_images(product_id: str, images: ProductImageUpdate, current_user: dict = Depends(require_roles([UserRole.ADMIN]))):
    """Update product images (admin only)"""
    product = await db.products.find_one({"id": product_id})
    if not product:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    
    update_data = {}
    if images.main_image is not None:
        update_data["main_image"] = images.main_image
    if images.images is not None:
        # Limit to 3 additional images
        update_data["images"] = images.images[:3]
    
    await db.products.update_one({"id": product_id}, {"$set": update_data})
    return {"message": "Imágenes actualizadas"}

@api_router.post("/products/{product_id}/upload-image")
async def upload_product_image(
    product_id: str,
    image_type: str = Form(...),  # "main" or "additional"
    image_index: int = Form(0),  # 0, 1, 2 for additional images
    file: UploadFile = File(...),
    current_user: dict = Depends(require_roles([UserRole.ADMIN]))
):
    """Upload a product image file"""
    product = await db.products.find_one({"id": product_id})
    if not product:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    
    # Validate file type
    allowed_types = ["image/jpeg", "image/png", "image/webp", "image/gif"]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Tipo de archivo no permitido. Use JPEG, PNG, WebP o GIF")
    
    # Generate unique filename
    ext = file.filename.split(".")[-1] if "." in file.filename else "jpg"
    filename = f"{product_id}_{image_type}_{image_index}_{uuid.uuid4().hex[:8]}.{ext}"
    file_path = UPLOADS_DIR / "products" / filename
    
    # Save file
    try:
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al guardar imagen: {str(e)}")
    
    # Generate URL path
    image_url = f"/api/uploads/products/{filename}"
    
    # Update product in database
    if image_type == "main":
        await db.products.update_one({"id": product_id}, {"$set": {"main_image": image_url}})
    else:
        # Get current images array
        current_images = product.get("images", ["", "", ""])
        # Ensure we have 3 slots
        while len(current_images) < 3:
            current_images.append("")
        # Update the specific index
        if 0 <= image_index < 3:
            current_images[image_index] = image_url
        await db.products.update_one({"id": product_id}, {"$set": {"images": current_images}})
    
    return {"message": "Imagen subida correctamente", "url": image_url}

@api_router.delete("/products/{product_id}/image")
async def delete_product_image(
    product_id: str,
    image_type: str,  # "main" or "additional"
    image_index: int = 0,
    current_user: dict = Depends(require_roles([UserRole.ADMIN]))
):
    """Delete a product image"""
    product = await db.products.find_one({"id": product_id})
    if not product:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    
    if image_type == "main":
        # Delete main image file if it's a local file
        if product.get("main_image") and product["main_image"].startswith("/api/uploads"):
            filename = product["main_image"].split("/")[-1]
            file_path = UPLOADS_DIR / "products" / filename
            if file_path.exists():
                file_path.unlink()
        await db.products.update_one({"id": product_id}, {"$set": {"main_image": None}})
    else:
        current_images = product.get("images", ["", "", ""])
        if 0 <= image_index < len(current_images):
            # Delete file if local
            if current_images[image_index] and current_images[image_index].startswith("/api/uploads"):
                filename = current_images[image_index].split("/")[-1]
                file_path = UPLOADS_DIR / "products" / filename
                if file_path.exists():
                    file_path.unlink()
            current_images[image_index] = ""
            await db.products.update_one({"id": product_id}, {"$set": {"images": current_images}})
    
    return {"message": "Imagen eliminada"}

@api_router.delete("/products/{product_id}")
async def delete_product(product_id: str, current_user: dict = Depends(require_roles([UserRole.ADMIN]))):
    result = await db.products.delete_one({"id": product_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    # Also delete associated recipe
    await db.recipes.delete_one({"product_id": product_id})
    return {"message": "Producto eliminado"}

# ============== CATALOG EXPORT ROUTES ==============

@api_router.get("/catalog/export/{cafeteria_id}")
async def export_catalog(cafeteria_id: str, format: str = "json", current_user: dict = Depends(require_roles([UserRole.ADMIN, UserRole.GERENTE]))):
    """Export product catalog for a specific cafeteria"""
    cafeteria = await db.cafeterias.find_one({"id": cafeteria_id}, {"_id": 0})
    if not cafeteria:
        raise HTTPException(status_code=404, detail="Cafetería no encontrada")
    
    products = await db.products.find({"is_active": True}, {"_id": 0}).to_list(1000)
    categories = {c["id"]: c["name"] for c in await db.categories.find({}, {"_id": 0}).to_list(100)}
    
    # Get inventory for this cafeteria
    inventory = {i["product_id"]: i["quantity"] for i in await db.inventory.find({"cafeteria_id": cafeteria_id}, {"_id": 0}).to_list(1000)}
    
    catalog = {
        "cafeteria": {
            "id": cafeteria["id"],
            "name": cafeteria["name"],
            "address": cafeteria["address"],
            "phone": cafeteria.get("phone")
        },
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "products": []
    }
    
    for p in products:
        product_data = {
            "id": p["id"],
            "name": p["name"],
            "description": p.get("description", ""),
            "category": categories.get(p["category_id"], "Sin categoría"),
            "price": p["price"],
            "cost": p["cost"],
            "main_image": p.get("main_image"),
            "images": p.get("images", []),
            "stock": inventory.get(p["id"], 0),
            "is_available": inventory.get(p["id"], 0) > 0
        }
        catalog["products"].append(product_data)
    
    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["ID", "Nombre", "Descripción", "Categoría", "Precio", "Costo", "Stock", "Disponible", "Imagen Principal"])
        for p in catalog["products"]:
            writer.writerow([
                p["id"], p["name"], p["description"], p["category"],
                p["price"], p["cost"], p["stock"], "Sí" if p["is_available"] else "No",
                p["main_image"] or ""
            ])
        
        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=catalogo_{cafeteria['name']}_{datetime.now().strftime('%Y%m%d')}.csv"}
        )
    
    return catalog

# ============== PDF/EXCEL REPORTS WITH LOGO ==============

async def get_tenant_logo_path(tenant_id: str) -> Optional[Path]:
    """Get the path to tenant's logo file"""
    if not tenant_id:
        return None
    logos_dir = UPLOADS_DIR / "logos"
    for logo_file in logos_dir.glob(f"logo_{tenant_id}.*"):
        return logo_file
    return None

@api_router.get("/reports/sales/pdf")
async def export_sales_pdf(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    cafeteria_id: Optional[str] = None,
    current_user: dict = Depends(require_roles([UserRole.ADMIN, UserRole.GERENTE]))
):
    """Export sales report as PDF with tenant logo"""
    tenant_id = current_user.get("tenant_id")
    tenant_filter = get_tenant_filter(current_user)
    
    # Get tenant info
    tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0}) if tenant_id else None
    business_name = tenant.get("business_name", "Reporte") if tenant else "Reporte"
    
    # Build query
    query = {**tenant_filter}
    if cafeteria_id:
        query["cafeteria_id"] = cafeteria_id
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date
        else:
            query["created_at"] = {"$lte": end_date}
    
    sales = await db.sales.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    cafeterias = {c["id"]: c["name"] for c in await db.cafeterias.find(tenant_filter, {"_id": 0}).to_list(100)}
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, alignment=TA_CENTER, textColor=colors.HexColor('#708238'))
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10, alignment=TA_CENTER, textColor=colors.gray)
    
    # Logo
    logo_path = await get_tenant_logo_path(tenant_id)
    if logo_path and logo_path.exists():
        try:
            logo = RLImage(str(logo_path), width=1.5*inch, height=1.5*inch)
            logo.hAlign = 'CENTER'
            elements.append(logo)
            elements.append(Spacer(1, 0.2*inch))
        except:
            pass
    
    # Header
    elements.append(Paragraph(business_name, title_style))
    elements.append(Paragraph("Reporte de Ventas", styles['Heading2']))
    
    date_range = ""
    if start_date and end_date:
        date_range = f"Del {start_date[:10]} al {end_date[:10]}"
    elif start_date:
        date_range = f"Desde {start_date[:10]}"
    elif end_date:
        date_range = f"Hasta {end_date[:10]}"
    else:
        date_range = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    
    elements.append(Paragraph(date_range, subtitle_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Summary
    total_sales = sum(s["total"] for s in sales)
    total_profit = sum(s["profit"] for s in sales)
    
    summary_data = [
        ["Total Ventas", f"${total_sales:,.2f}"],
        ["Total Utilidad", f"${total_profit:,.2f}"],
        ["Número de Transacciones", str(len(sales))]
    ]
    summary_table = Table(summary_data, colWidths=[2.5*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f5f5f5')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#333333')),
        ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#708238')),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('RIGHTPADDING', (0, 0), (-1, -1), 12),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Sales table
    elements.append(Paragraph("Detalle de Ventas", styles['Heading3']))
    elements.append(Spacer(1, 0.1*inch))
    
    table_data = [["Fecha", "Cafetería", "Productos", "Total", "Utilidad", "Método"]]
    for s in sales[:50]:  # Limit to 50 for PDF
        items_str = ", ".join([f"{i['product_name']} x{i['quantity']}" for i in s.get("items", [])[:2]])
        if len(s.get("items", [])) > 2:
            items_str += f" (+{len(s['items'])-2})"
        table_data.append([
            s["created_at"][:10] if s.get("created_at") else "",
            cafeterias.get(s.get("cafeteria_id"), "")[:15],
            items_str[:30],
            f"${s['total']:,.2f}",
            f"${s['profit']:,.2f}",
            s.get("payment_method", "")[:10]
        ])
    
    sales_table = Table(table_data, colWidths=[0.9*inch, 1.1*inch, 2.2*inch, 0.9*inch, 0.9*inch, 0.8*inch])
    sales_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#708238')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (3, 0), (4, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e0e0e0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')]),
    ]))
    elements.append(sales_table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"ventas_{business_name}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/reports/sales/excel")
async def export_sales_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    cafeteria_id: Optional[str] = None,
    current_user: dict = Depends(require_roles([UserRole.ADMIN, UserRole.GERENTE]))
):
    """Export sales report as Excel with tenant logo"""
    tenant_id = current_user.get("tenant_id")
    tenant_filter = get_tenant_filter(current_user)
    
    tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0}) if tenant_id else None
    business_name = tenant.get("business_name", "Reporte") if tenant else "Reporte"
    
    query = {**tenant_filter}
    if cafeteria_id:
        query["cafeteria_id"] = cafeteria_id
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date
        else:
            query["created_at"] = {"$lte": end_date}
    
    sales = await db.sales.find(query, {"_id": 0}).sort("created_at", -1).to_list(10000)
    cafeterias = {c["id"]: c["name"] for c in await db.cafeterias.find(tenant_filter, {"_id": 0}).to_list(100)}
    
    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"
    
    # Styles
    header_fill = PatternFill(start_color="708238", end_color="708238", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=16, color="708238")
    money_font = Font(color="708238")
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )
    
    # Logo
    logo_path = await get_tenant_logo_path(tenant_id)
    start_row = 1
    if logo_path and logo_path.exists():
        try:
            img = XLImage(str(logo_path))
            img.width = 80
            img.height = 80
            ws.add_image(img, 'A1')
            start_row = 6
        except:
            pass
    
    # Title
    ws.merge_cells(f'A{start_row}:G{start_row}')
    ws[f'A{start_row}'] = business_name
    ws[f'A{start_row}'].font = title_font
    ws[f'A{start_row}'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells(f'A{start_row+1}:G{start_row+1}')
    ws[f'A{start_row+1}'] = "Reporte de Ventas"
    ws[f'A{start_row+1}'].font = Font(bold=True, size=12)
    ws[f'A{start_row+1}'].alignment = Alignment(horizontal='center')
    
    # Summary
    total_sales = sum(s["total"] for s in sales)
    total_profit = sum(s["profit"] for s in sales)
    
    summary_row = start_row + 3
    ws[f'A{summary_row}'] = "Total Ventas:"
    ws[f'B{summary_row}'] = f"${total_sales:,.2f}"
    ws[f'B{summary_row}'].font = money_font
    ws[f'C{summary_row}'] = "Total Utilidad:"
    ws[f'D{summary_row}'] = f"${total_profit:,.2f}"
    ws[f'D{summary_row}'].font = money_font
    ws[f'E{summary_row}'] = "Transacciones:"
    ws[f'F{summary_row}'] = len(sales)
    
    # Headers
    header_row = summary_row + 2
    headers = ["Fecha", "Cafetería", "Productos", "Subtotal", "Impuesto", "Total", "Utilidad", "Método Pago"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Data
    for row_idx, sale in enumerate(sales, header_row + 1):
        items_str = ", ".join([f"{i['product_name']} x{i['quantity']}" for i in sale.get("items", [])])
        
        ws.cell(row=row_idx, column=1, value=sale.get("created_at", "")[:10]).border = thin_border
        ws.cell(row=row_idx, column=2, value=cafeterias.get(sale.get("cafeteria_id"), "")).border = thin_border
        ws.cell(row=row_idx, column=3, value=items_str).border = thin_border
        ws.cell(row=row_idx, column=4, value=sale.get("subtotal", 0)).border = thin_border
        ws.cell(row=row_idx, column=5, value=sale.get("tax", 0)).border = thin_border
        ws.cell(row=row_idx, column=6, value=sale.get("total", 0)).border = thin_border
        ws.cell(row=row_idx, column=7, value=sale.get("profit", 0)).border = thin_border
        ws.cell(row=row_idx, column=8, value=sale.get("payment_method", "")).border = thin_border
        
        # Format money columns
        for col in [4, 5, 6, 7]:
            ws.cell(row=row_idx, column=col).number_format = '$#,##0.00'
    
    # Auto-width columns
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 15
    ws.column_dimensions['C'].width = 40
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"ventas_{business_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/reports/products/pdf")
async def export_products_pdf(
    cafeteria_id: Optional[str] = None,
    current_user: dict = Depends(require_roles([UserRole.ADMIN, UserRole.GERENTE]))
):
    """Export products catalog as PDF with tenant logo"""
    tenant_id = current_user.get("tenant_id")
    tenant_filter = get_tenant_filter(current_user)
    
    tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0}) if tenant_id else None
    business_name = tenant.get("business_name", "Catálogo") if tenant else "Catálogo"
    
    products = await db.products.find({**tenant_filter, "is_active": True}, {"_id": 0}).to_list(1000)
    categories = {c["id"]: c["name"] for c in await db.categories.find(tenant_filter, {"_id": 0}).to_list(100)}
    
    inventory = {}
    if cafeteria_id:
        inv_items = await db.inventory.find({"cafeteria_id": cafeteria_id}, {"_id": 0}).to_list(1000)
        inventory = {i["product_id"]: i["quantity"] for i in inv_items}
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, alignment=TA_CENTER, textColor=colors.HexColor('#708238'))
    
    logo_path = await get_tenant_logo_path(tenant_id)
    if logo_path and logo_path.exists():
        try:
            logo = RLImage(str(logo_path), width=1.5*inch, height=1.5*inch)
            logo.hAlign = 'CENTER'
            elements.append(logo)
            elements.append(Spacer(1, 0.2*inch))
        except:
            pass
    
    elements.append(Paragraph(business_name, title_style))
    elements.append(Paragraph("Catálogo de Productos", styles['Heading2']))
    elements.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", 
                              ParagraphStyle('Date', alignment=TA_CENTER, textColor=colors.gray, fontSize=10)))
    elements.append(Spacer(1, 0.3*inch))
    
    table_data = [["Producto", "Categoría", "Precio", "Costo", "Margen", "Stock"]]
    for p in products:
        margin = ((p["price"] - p["cost"]) / p["price"] * 100) if p["price"] > 0 else 0
        stock = inventory.get(p["id"], "-") if inventory else "-"
        table_data.append([
            p["name"][:25],
            categories.get(p.get("category_id"), "Sin categoría")[:15],
            f"${p['price']:,.2f}",
            f"${p['cost']:,.2f}",
            f"{margin:.1f}%",
            str(stock)
        ])
    
    products_table = Table(table_data, colWidths=[2.2*inch, 1.3*inch, 0.9*inch, 0.9*inch, 0.8*inch, 0.7*inch])
    products_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#708238')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e0e0e0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')]),
    ]))
    elements.append(products_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"catalogo_{business_name}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ============== INVENTORY ROUTES ==============

@api_router.get("/inventory", response_model=List[InventoryItemResponse])
async def get_inventory(cafeteria_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {}
    if cafeteria_id:
        query["cafeteria_id"] = cafeteria_id
    elif current_user["role"] == UserRole.GERENTE and current_user.get("cafeteria_id"):
        query["cafeteria_id"] = current_user["cafeteria_id"]
    
    inventory = await db.inventory.find(query, {"_id": 0}).to_list(1000)
    
    products = {p["id"]: p["name"] for p in await db.products.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(1000)}
    cafeterias = {c["id"]: c["name"] for c in await db.cafeterias.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(100)}
    
    result = []
    for item in inventory:
        item["product_name"] = products.get(item["product_id"], "Desconocido")
        item["cafeteria_name"] = cafeterias.get(item["cafeteria_id"], "Desconocida")
        item["is_low_stock"] = item["quantity"] <= item["min_stock"]
        result.append(InventoryItemResponse(**item))
    
    return result

@api_router.post("/inventory", response_model=InventoryItemResponse)
async def create_inventory_item(item: InventoryItemCreate, current_user: dict = Depends(require_roles([UserRole.ADMIN, UserRole.GERENTE]))):
    existing = await db.inventory.find_one({
        "product_id": item.product_id,
        "cafeteria_id": item.cafeteria_id
    })
    if existing:
        raise HTTPException(status_code=400, detail="Ya existe inventario para este producto en esta cafetería")
    
    item_id = str(uuid.uuid4())
    item_dict = item.model_dump()
    item_dict["id"] = item_id
    item_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.inventory.insert_one(item_dict)
    
    product = await db.products.find_one({"id": item.product_id}, {"_id": 0, "name": 1})
    cafeteria = await db.cafeterias.find_one({"id": item.cafeteria_id}, {"_id": 0, "name": 1})
    
    return InventoryItemResponse(
        **item_dict,
        product_name=product["name"] if product else "Desconocido",
        cafeteria_name=cafeteria["name"] if cafeteria else "Desconocida",
        is_low_stock=item.quantity <= item.min_stock
    )

@api_router.post("/inventory/movement")
async def record_inventory_movement(movement: InventoryMovement, current_user: dict = Depends(require_roles([UserRole.ADMIN, UserRole.GERENTE]))):
    inventory_item = await db.inventory.find_one({"id": movement.inventory_id})
    if not inventory_item:
        raise HTTPException(status_code=404, detail="Item de inventario no encontrado")
    
    new_quantity = inventory_item["quantity"]
    if movement.movement_type in ["entrada"]:
        new_quantity += movement.quantity
    elif movement.movement_type in ["salida", "merma"]:
        new_quantity -= movement.quantity
    else:
        new_quantity = movement.quantity
    
    if new_quantity < 0:
        raise HTTPException(status_code=400, detail="No hay suficiente stock")
    
    await db.inventory.update_one({"id": movement.inventory_id}, {"$set": {"quantity": new_quantity}})
    
    movement_log = {
        "id": str(uuid.uuid4()),
        "inventory_id": movement.inventory_id,
        "quantity": movement.quantity,
        "movement_type": movement.movement_type,
        "reason": movement.reason,
        "previous_quantity": inventory_item["quantity"],
        "new_quantity": new_quantity,
        "created_by": current_user["user_id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.inventory_movements.insert_one(movement_log)
    
    return {"message": "Movimiento registrado", "new_quantity": new_quantity}

@api_router.get("/inventory/movements")
async def get_inventory_movements(inventory_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {}
    if inventory_id:
        query["inventory_id"] = inventory_id
    
    movements = await db.inventory_movements.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return movements

# ============== SUPPLIER ROUTES ==============

@api_router.get("/suppliers", response_model=List[SupplierResponse])
async def get_suppliers(current_user: dict = Depends(get_current_user)):
    suppliers = await db.suppliers.find({}, {"_id": 0}).to_list(100)
    return [SupplierResponse(**s) for s in suppliers]

@api_router.post("/suppliers", response_model=SupplierResponse)
async def create_supplier(supplier: SupplierCreate, current_user: dict = Depends(require_roles([UserRole.ADMIN, UserRole.GERENTE]))):
    supplier_id = str(uuid.uuid4())
    supplier_dict = supplier.model_dump()
    supplier_dict["id"] = supplier_id
    supplier_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.suppliers.insert_one(supplier_dict)
    return SupplierResponse(**supplier_dict)

@api_router.put("/suppliers/{supplier_id}", response_model=SupplierResponse)
async def update_supplier(supplier_id: str, supplier: SupplierBase, current_user: dict = Depends(require_roles([UserRole.ADMIN, UserRole.GERENTE]))):
    result = await db.suppliers.update_one({"id": supplier_id}, {"$set": supplier.model_dump()})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")
    updated = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    return SupplierResponse(**updated)

@api_router.delete("/suppliers/{supplier_id}")
async def delete_supplier(supplier_id: str, current_user: dict = Depends(require_roles([UserRole.ADMIN]))):
    result = await db.suppliers.delete_one({"id": supplier_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")
    return {"message": "Proveedor eliminado"}

# ============== PURCHASE ROUTES (UPDATED for ingredients) ==============

@api_router.get("/purchases", response_model=List[PurchaseResponse])
async def get_purchases(cafeteria_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {}
    if cafeteria_id:
        query["cafeteria_id"] = cafeteria_id
    elif current_user["role"] == UserRole.GERENTE and current_user.get("cafeteria_id"):
        query["cafeteria_id"] = current_user["cafeteria_id"]
    
    purchases = await db.purchases.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    
    suppliers = {s["id"]: s["name"] for s in await db.suppliers.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(100)}
    cafeterias = {c["id"]: c["name"] for c in await db.cafeterias.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(100)}
    
    result = []
    for p in purchases:
        p["supplier_name"] = suppliers.get(p["supplier_id"], "Desconocido")
        p["cafeteria_name"] = cafeterias.get(p["cafeteria_id"], "Desconocida")
        result.append(PurchaseResponse(**p))
    
    return result

@api_router.post("/purchases", response_model=PurchaseResponse)
async def create_purchase(purchase: PurchaseCreate, current_user: dict = Depends(require_roles([UserRole.ADMIN, UserRole.GERENTE]))):
    purchase_id = str(uuid.uuid4())
    
    total = 0
    enriched_items = []
    products = {p["id"]: p for p in await db.products.find({}, {"_id": 0}).to_list(1000)}
    ingredients = {i["id"]: i for i in await db.ingredients.find({}, {"_id": 0}).to_list(1000)}
    
    for item in purchase.items:
        item_total = item.quantity * item.unit_cost
        total += item_total
        
        item_name = "Desconocido"
        item_type = "unknown"
        
        if item.ingredient_id:
            ingredient = ingredients.get(item.ingredient_id)
            item_name = ingredient["name"] if ingredient else "Desconocido"
            item_type = "ingredient"
            
            # Update ingredient inventory
            inv_item = await db.ingredient_inventory.find_one({
                "ingredient_id": item.ingredient_id,
                "cafeteria_id": purchase.cafeteria_id
            })
            if inv_item:
                await db.ingredient_inventory.update_one(
                    {"id": inv_item["id"]},
                    {"$inc": {"quantity": item.quantity}}
                )
            else:
                # Create new inventory item
                await db.ingredient_inventory.insert_one({
                    "id": str(uuid.uuid4()),
                    "ingredient_id": item.ingredient_id,
                    "cafeteria_id": purchase.cafeteria_id,
                    "quantity": item.quantity,
                    "min_stock": 10.0,
                    "created_at": datetime.now(timezone.utc).isoformat()
                })
            
            # Update ingredient cost if different
            if ingredient and item.unit_cost != ingredient.get("cost_per_unit", 0):
                await db.ingredients.update_one(
                    {"id": item.ingredient_id},
                    {"$set": {"cost_per_unit": item.unit_cost}}
                )
        
        elif item.product_id:
            product = products.get(item.product_id)
            item_name = product["name"] if product else "Desconocido"
            item_type = "product"
            
            # Update product inventory
            inv_item = await db.inventory.find_one({
                "product_id": item.product_id,
                "cafeteria_id": purchase.cafeteria_id
            })
            if inv_item:
                await db.inventory.update_one(
                    {"id": inv_item["id"]},
                    {"$inc": {"quantity": item.quantity}}
                )
        
        enriched_items.append({
            "ingredient_id": item.ingredient_id,
            "product_id": item.product_id,
            "item_name": item_name,
            "item_type": item_type,
            "quantity": item.quantity,
            "unit_cost": item.unit_cost,
            "total": item_total
        })
    
    purchase_dict = {
        "id": purchase_id,
        "supplier_id": purchase.supplier_id,
        "cafeteria_id": purchase.cafeteria_id,
        "items": enriched_items,
        "total": total,
        "notes": purchase.notes,
        "created_by": current_user["user_id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.purchases.insert_one(purchase_dict)
    
    supplier = await db.suppliers.find_one({"id": purchase.supplier_id}, {"_id": 0, "name": 1})
    cafeteria = await db.cafeterias.find_one({"id": purchase.cafeteria_id}, {"_id": 0, "name": 1})
    
    return PurchaseResponse(
        **purchase_dict,
        supplier_name=supplier["name"] if supplier else "Desconocido",
        cafeteria_name=cafeteria["name"] if cafeteria else "Desconocida"
    )

# ============== SALES ROUTES (UPDATED with ingredient deduction) ==============

@api_router.get("/sales", response_model=List[SaleResponse])
async def get_sales(
    cafeteria_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    tenant_filter = get_tenant_filter(current_user)
    query = {**tenant_filter}
    if cafeteria_id:
        query["cafeteria_id"] = cafeteria_id
    elif current_user["role"] in [UserRole.GERENTE, UserRole.CAJERO] and current_user.get("cafeteria_id"):
        query["cafeteria_id"] = current_user["cafeteria_id"]
    
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date
        else:
            query["created_at"] = {"$lte": end_date}
    
    sales = await db.sales.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    cafeterias = {c["id"]: c["name"] for c in await db.cafeterias.find(tenant_filter, {"_id": 0, "id": 1, "name": 1}).to_list(100)}
    
    result = []
    for s in sales:
        s["cafeteria_name"] = cafeterias.get(s["cafeteria_id"], "Desconocida")
        result.append(SaleResponse(**s))
    
    return result

@api_router.post("/sales", response_model=SaleResponse)
async def create_sale(sale: SaleCreate, current_user: dict = Depends(get_current_user)):
    tenant_filter = get_tenant_filter(current_user)
    sale_id = str(uuid.uuid4())
    
    subtotal = sum(item.subtotal for item in sale.items)
    tax = subtotal * 0.16
    total = subtotal + tax
    
    products = {p["id"]: p for p in await db.products.find(tenant_filter, {"_id": 0}).to_list(1000)}
    recipes = {r["product_id"]: r for r in await db.recipes.find(tenant_filter, {"_id": 0}).to_list(1000)}
    ingredients = {i["id"]: i for i in await db.ingredients.find(tenant_filter, {"_id": 0}).to_list(1000)}
    
    cost_total = 0
    enriched_items = []
    
    for item in sale.items:
        product = products.get(item.product_id)
        item_cost = (product["cost"] if product else 0) * item.quantity
        cost_total += item_cost
        
        enriched_items.append({
            "product_id": item.product_id,
            "product_name": item.product_name,
            "quantity": item.quantity,
            "unit_price": item.unit_price,
            "subtotal": item.subtotal,
            "cost": item_cost
        })
        
        # Update product inventory
        await db.inventory.update_one(
            {"product_id": item.product_id, "cafeteria_id": sale.cafeteria_id},
            {"$inc": {"quantity": -item.quantity}}
        )
        
        # Deduct ingredients if recipe exists and auto_deduct is enabled
        recipe = recipes.get(item.product_id)
        if recipe and recipe.get("auto_deduct", True):
            for ing_item in recipe.get("ingredients", []):
                # Calculate quantity to deduct (per portion * quantity sold)
                qty_to_deduct = (ing_item["quantity"] / recipe.get("portions", 1)) * item.quantity
                
                # Find ingredient inventory for this cafeteria
                ing_inv = await db.ingredient_inventory.find_one({
                    "ingredient_id": ing_item["ingredient_id"],
                    "cafeteria_id": sale.cafeteria_id
                })
                
                if ing_inv:
                    # Deduct from inventory
                    new_qty = max(0, ing_inv["quantity"] - qty_to_deduct)
                    await db.ingredient_inventory.update_one(
                        {"id": ing_inv["id"]},
                        {"$set": {"quantity": new_qty}}
                    )
                    
                    # Log the movement
                    await db.ingredient_movements.insert_one({
                        "id": str(uuid.uuid4()),
                        "inventory_id": ing_inv["id"],
                        "ingredient_id": ing_item["ingredient_id"],
                        "cafeteria_id": sale.cafeteria_id,
                        "quantity": qty_to_deduct,
                        "movement_type": "consumo_venta",
                        "reason": f"Venta de {item.quantity}x {item.product_name}",
                        "sale_id": sale_id,
                        "previous_quantity": ing_inv["quantity"],
                        "new_quantity": new_qty,
                        "created_by": current_user["user_id"],
                        "created_at": datetime.now(timezone.utc).isoformat()
                    })
    
    profit = subtotal - cost_total
    
    sale_dict = {
        "id": sale_id,
        "cafeteria_id": sale.cafeteria_id,
        "tenant_id": current_user.get("tenant_id"),
        "items": enriched_items,
        "subtotal": round(subtotal, 2),
        "tax": round(tax, 2),
        "total": round(total, 2),
        "cost_total": round(cost_total, 2),
        "profit": round(profit, 2),
        "payment_method": sale.payment_method,
        "clip_transaction_id": sale.clip_transaction_id,
        "notes": sale.notes,
        "created_by": current_user["user_id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.sales.insert_one(sale_dict)
    
    cafeteria = await db.cafeterias.find_one({"id": sale.cafeteria_id}, {"_id": 0, "name": 1})
    
    return SaleResponse(**sale_dict, cafeteria_name=cafeteria["name"] if cafeteria else "Desconocida")

# ============== DASHBOARD / REPORTS ROUTES ==============

@api_router.get("/dashboard/stats", response_model=DashboardStats)
async def get_dashboard_stats(cafeteria_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    tenant_filter = get_tenant_filter(current_user)
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    month_start = today.replace(day=1)
    
    query = {**tenant_filter}
    if cafeteria_id:
        query["cafeteria_id"] = cafeteria_id
    elif current_user["role"] in [UserRole.GERENTE, UserRole.CAJERO] and current_user.get("cafeteria_id"):
        query["cafeteria_id"] = current_user["cafeteria_id"]
    
    all_sales = await db.sales.find(query, {"_id": 0}).to_list(10000)
    
    today_str = today.isoformat()
    month_str = month_start.isoformat()
    
    today_sales = [s for s in all_sales if s["created_at"] >= today_str]
    month_sales = [s for s in all_sales if s["created_at"] >= month_str]
    
    total_sales_today = sum(s["total"] for s in today_sales)
    total_sales_month = sum(s["total"] for s in month_sales)
    total_profit_today = sum(s["profit"] for s in today_sales)
    total_profit_month = sum(s["profit"] for s in month_sales)
    
    # Product inventory alerts
    inv_query = {**tenant_filter}
    if cafeteria_id:
        inv_query["cafeteria_id"] = cafeteria_id
    
    inventory = await db.inventory.find(inv_query, {"_id": 0}).to_list(1000)
    low_stock_alerts = sum(1 for i in inventory if i["quantity"] <= i["min_stock"])
    
    # Ingredient inventory alerts
    ing_inventory = await db.ingredient_inventory.find(inv_query, {"_id": 0}).to_list(1000)
    low_ingredient_alerts = sum(1 for i in ing_inventory if i["quantity"] <= i["min_stock"])
    
    # Top products
    product_sales = {}
    for sale in month_sales:
        for item in sale["items"]:
            pid = item["product_id"]
            if pid not in product_sales:
                product_sales[pid] = {"name": item["product_name"], "quantity": 0, "revenue": 0}
            product_sales[pid]["quantity"] += item["quantity"]
            product_sales[pid]["revenue"] += item["subtotal"]
    
    top_products = sorted(product_sales.values(), key=lambda x: x["revenue"], reverse=True)[:5]
    
    # Sales by cafeteria
    cafeterias = {c["id"]: c["name"] for c in await db.cafeterias.find(tenant_filter, {"_id": 0, "id": 1, "name": 1}).to_list(100)}
    cafe_sales = {}
    for sale in month_sales:
        cid = sale["cafeteria_id"]
        if cid not in cafe_sales:
            cafe_sales[cid] = {"name": cafeterias.get(cid, "Desconocida"), "total": 0, "profit": 0}
        cafe_sales[cid]["total"] += sale["total"]
        cafe_sales[cid]["profit"] += sale["profit"]
    
    sales_by_cafeteria = list(cafe_sales.values())
    
    # Sales trend
    sales_trend = []
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        day_str = day.isoformat()
        next_day_str = (day + timedelta(days=1)).isoformat()
        day_sales = [s for s in all_sales if day_str <= s["created_at"] < next_day_str]
        sales_trend.append({
            "date": day.strftime("%Y-%m-%d"),
            "day": day.strftime("%a"),
            "total": sum(s["total"] for s in day_sales),
            "profit": sum(s["profit"] for s in day_sales)
        })
    
    return DashboardStats(
        total_sales_today=round(total_sales_today, 2),
        total_sales_month=round(total_sales_month, 2),
        total_profit_today=round(total_profit_today, 2),
        total_profit_month=round(total_profit_month, 2),
        sales_count_today=len(today_sales),
        low_stock_alerts=low_stock_alerts,
        low_ingredient_alerts=low_ingredient_alerts,
        top_products=top_products,
        sales_by_cafeteria=sales_by_cafeteria,
        sales_trend=sales_trend
    )

@api_router.get("/reports/sales-comparison")
async def get_sales_comparison(current_user: dict = Depends(require_roles([UserRole.ADMIN, UserRole.GERENTE]))):
    tenant_filter = get_tenant_filter(current_user)
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    month_start = today.replace(day=1)
    month_str = month_start.isoformat()
    
    sales = await db.sales.find({"created_at": {"$gte": month_str}}, {"_id": 0}).to_list(10000)
    cafeterias = await db.cafeterias.find({}, {"_id": 0}).to_list(100)
    
    comparison = []
    for cafe in cafeterias:
        cafe_sales = [s for s in sales if s["cafeteria_id"] == cafe["id"]]
        comparison.append({
            "cafeteria_id": cafe["id"],
            "cafeteria_name": cafe["name"],
            "total_sales": round(sum(s["total"] for s in cafe_sales), 2),
            "total_profit": round(sum(s["profit"] for s in cafe_sales), 2),
            "transaction_count": len(cafe_sales),
            "average_ticket": round(sum(s["total"] for s in cafe_sales) / len(cafe_sales), 2) if cafe_sales else 0
        })
    
    return comparison

@api_router.get("/reports/profit-analysis")
async def get_profit_analysis(cafeteria_id: Optional[str] = None, current_user: dict = Depends(require_roles([UserRole.ADMIN, UserRole.GERENTE]))):
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    month_start = today.replace(day=1)
    month_str = month_start.isoformat()
    
    query = {"created_at": {"$gte": month_str}}
    if cafeteria_id:
        query["cafeteria_id"] = cafeteria_id
    
    sales = await db.sales.find(query, {"_id": 0}).to_list(10000)
    purchases = await db.purchases.find(query, {"_id": 0}).to_list(1000)
    
    total_revenue = sum(s["subtotal"] for s in sales)
    total_cost_of_goods = sum(s["cost_total"] for s in sales)
    total_purchases = sum(p["total"] for p in purchases)
    gross_profit = total_revenue - total_cost_of_goods
    
    return {
        "total_revenue": round(total_revenue, 2),
        "total_cost_of_goods": round(total_cost_of_goods, 2),
        "gross_profit": round(gross_profit, 2),
        "gross_margin_percent": round((gross_profit / total_revenue * 100) if total_revenue > 0 else 0, 2),
        "total_purchases": round(total_purchases, 2),
        "transaction_count": len(sales)
    }

@api_router.get("/reports/ingredient-consumption")
async def get_ingredient_consumption(cafeteria_id: Optional[str] = None, days: int = 30, current_user: dict = Depends(require_roles([UserRole.ADMIN, UserRole.GERENTE]))):
    """Report of ingredient consumption over time"""
    start_date = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
    
    query = {"movement_type": "consumo_venta", "created_at": {"$gte": start_date}}
    if cafeteria_id:
        query["cafeteria_id"] = cafeteria_id
    
    movements = await db.ingredient_movements.find(query, {"_id": 0}).to_list(10000)
    ingredients = {i["id"]: i for i in await db.ingredients.find({}, {"_id": 0}).to_list(1000)}
    
    consumption = {}
    for mov in movements:
        ing_id = mov["ingredient_id"]
        if ing_id not in consumption:
            ing = ingredients.get(ing_id, {})
            consumption[ing_id] = {
                "ingredient_id": ing_id,
                "ingredient_name": ing.get("name", "Desconocido"),
                "unit": ing.get("unit", "unidad"),
                "total_consumed": 0,
                "total_cost": 0
            }
        consumption[ing_id]["total_consumed"] += mov["quantity"]
        consumption[ing_id]["total_cost"] += mov["quantity"] * ingredients.get(ing_id, {}).get("cost_per_unit", 0)
    
    result = list(consumption.values())
    for item in result:
        item["total_consumed"] = round(item["total_consumed"], 2)
        item["total_cost"] = round(item["total_cost"], 2)
    
    return sorted(result, key=lambda x: x["total_cost"], reverse=True)

@api_router.get("/reports/theoretical-vs-actual")
async def get_theoretical_vs_actual(cafeteria_id: str, current_user: dict = Depends(require_roles([UserRole.ADMIN, UserRole.GERENTE]))):
    """Compare theoretical inventory (based on sales) vs actual inventory"""
    ingredients = {i["id"]: i for i in await db.ingredients.find({}, {"_id": 0}).to_list(1000)}
    
    # Get current inventory
    actual_inventory = await db.ingredient_inventory.find({"cafeteria_id": cafeteria_id}, {"_id": 0}).to_list(1000)
    
    # Calculate theoretical consumption from all sales
    all_movements = await db.ingredient_movements.find({
        "cafeteria_id": cafeteria_id,
        "movement_type": "consumo_venta"
    }, {"_id": 0}).to_list(50000)
    
    theoretical_consumed = {}
    for mov in all_movements:
        ing_id = mov["ingredient_id"]
        if ing_id not in theoretical_consumed:
            theoretical_consumed[ing_id] = 0
        theoretical_consumed[ing_id] += mov["quantity"]
    
    # Compare
    comparison = []
    for inv_item in actual_inventory:
        ing = ingredients.get(inv_item["ingredient_id"], {})
        theoretical = theoretical_consumed.get(inv_item["ingredient_id"], 0)
        
        # Get all entries (purchases)
        entries = await db.ingredient_movements.find({
            "cafeteria_id": cafeteria_id,
            "ingredient_id": inv_item["ingredient_id"],
            "movement_type": "entrada"
        }, {"_id": 0}).to_list(1000)
        total_entries = sum(e["quantity"] for e in entries)
        
        theoretical_remaining = total_entries - theoretical
        actual_remaining = inv_item["quantity"]
        variance = actual_remaining - theoretical_remaining
        
        comparison.append({
            "ingredient_id": inv_item["ingredient_id"],
            "ingredient_name": ing.get("name", "Desconocido"),
            "unit": ing.get("unit", "unidad"),
            "total_purchased": round(total_entries, 2),
            "theoretical_consumed": round(theoretical, 2),
            "theoretical_remaining": round(theoretical_remaining, 2),
            "actual_remaining": round(actual_remaining, 2),
            "variance": round(variance, 2),
            "variance_percent": round((variance / theoretical_remaining * 100) if theoretical_remaining > 0 else 0, 2)
        })
    
    return comparison

# ============== CLIP INTEGRATION (MOCK) ==============

@api_router.post("/clip/sync")
async def sync_clip_transactions(current_user: dict = Depends(require_roles([UserRole.ADMIN, UserRole.GERENTE]))):
    return {
        "message": "Sincronización con Clip completada (MOCK)",
        "note": "Para integración real, configure las credenciales de Clip en developer.clip.mx",
        "transactions_synced": 0
    }

@api_router.get("/clip/status")
async def get_clip_status(current_user: dict = Depends(get_current_user)):
    clip_api_key = os.environ.get("CLIP_API_KEY")
    return {
        "connected": bool(clip_api_key),
        "message": "Clip no configurado. Visite developer.clip.mx para obtener credenciales." if not clip_api_key else "Clip conectado"
    }

# ============== WHATSAPP ALERTS ==============

class WhatsAppNumberCreate(BaseModel):
    phone_number: str  # Format: +525591985187
    name: str
    is_active: bool = True

class WhatsAppNumberResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    phone_number: str
    name: str
    is_active: bool
    created_at: str

@api_router.get("/whatsapp/numbers", response_model=List[WhatsAppNumberResponse])
async def get_whatsapp_numbers(current_user: dict = Depends(require_roles([UserRole.ADMIN]))):
    """Get all WhatsApp numbers configured for alerts"""
    numbers = await db.whatsapp_numbers.find({}, {"_id": 0}).to_list(50)
    return [WhatsAppNumberResponse(**n) for n in numbers]

@api_router.post("/whatsapp/numbers", response_model=WhatsAppNumberResponse)
async def add_whatsapp_number(number: WhatsAppNumberCreate, current_user: dict = Depends(require_roles([UserRole.ADMIN]))):
    """Add a WhatsApp number for alerts"""
    # Validate format
    if not number.phone_number.startswith("+"):
        raise HTTPException(status_code=400, detail="El número debe empezar con + y código de país (ej: +525591985187)")
    
    # Check if exists
    existing = await db.whatsapp_numbers.find_one({"phone_number": number.phone_number})
    if existing:
        raise HTTPException(status_code=400, detail="Este número ya está registrado")
    
    number_id = str(uuid.uuid4())
    number_dict = number.model_dump()
    number_dict["id"] = number_id
    number_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.whatsapp_numbers.insert_one(number_dict)
    return WhatsAppNumberResponse(**number_dict)

@api_router.delete("/whatsapp/numbers/{number_id}")
async def delete_whatsapp_number(number_id: str, current_user: dict = Depends(require_roles([UserRole.ADMIN]))):
    """Remove a WhatsApp number"""
    result = await db.whatsapp_numbers.delete_one({"id": number_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Número no encontrado")
    return {"message": "Número eliminado"}

@api_router.get("/whatsapp/status")
async def get_whatsapp_status(current_user: dict = Depends(require_roles([UserRole.ADMIN]))):
    """Check WhatsApp integration status"""
    return {
        "configured": bool(twilio_client),
        "twilio_number": TWILIO_WHATSAPP_NUMBER,
        "message": "WhatsApp configurado correctamente" if twilio_client else "Twilio no configurado"
    }

@api_router.post("/whatsapp/send-test")
async def send_test_whatsapp(current_user: dict = Depends(require_roles([UserRole.ADMIN]))):
    """Send a test message to all configured numbers"""
    if not twilio_client:
        raise HTTPException(status_code=400, detail="Twilio no está configurado")
    
    numbers = await db.whatsapp_numbers.find({"is_active": True}, {"_id": 0}).to_list(50)
    if not numbers:
        raise HTTPException(status_code=400, detail="No hay números configurados para alertas")
    
    results = []
    for num in numbers:
        try:
            message = twilio_client.messages.create(
                body=f"🧪 Mensaje de prueba de Doré\n\nEste es un mensaje de prueba para verificar que las alertas de WhatsApp funcionan correctamente.\n\n✅ Conexión exitosa",
                from_=f"whatsapp:{TWILIO_WHATSAPP_NUMBER}",
                to=f"whatsapp:{num['phone_number']}"
            )
            results.append({"phone": num["phone_number"], "status": "sent", "sid": message.sid})
            logger.info(f"Test WhatsApp sent to {num['phone_number']}: {message.sid}")
        except Exception as e:
            results.append({"phone": num["phone_number"], "status": "failed", "error": str(e)})
            logger.error(f"Failed to send WhatsApp to {num['phone_number']}: {e}")
    
    return {"results": results}

@api_router.post("/whatsapp/send-alerts")
async def send_ingredient_alerts(current_user: dict = Depends(require_roles([UserRole.ADMIN, UserRole.GERENTE]))):
    """Send critical ingredient alerts to all configured WhatsApp numbers"""
    if not twilio_client:
        raise HTTPException(status_code=400, detail="Twilio no está configurado")
    
    # Get critical alerts (less than 3 days of stock)
    alerts = await get_critical_ingredient_alerts()
    
    if not alerts:
        return {"message": "No hay alertas críticas de ingredientes", "alerts_sent": 0}
    
    numbers = await db.whatsapp_numbers.find({"is_active": True}, {"_id": 0}).to_list(50)
    if not numbers:
        raise HTTPException(status_code=400, detail="No hay números configurados para alertas")
    
    # Build alert message
    message_lines = ["🚨 *ALERTA CRÍTICA - DORÉ*\n", "Ingredientes con stock crítico (< 3 días):\n"]
    
    for alert in alerts[:10]:  # Max 10 alerts per message
        message_lines.append(
            f"⚠️ *{alert['ingredient_name']}*\n"
            f"   📍 {alert['cafeteria_name']}\n"
            f"   📦 Stock: {alert['current_stock']:.1f} {alert['unit']}\n"
            f"   ⏱️ Días restantes: {alert['days_until_stockout']}\n"
            f"   🛒 Sugerido pedir: {alert['suggested_order']:.1f} {alert['unit']}\n"
        )
    
    message_lines.append(f"\n📅 {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    message_body = "\n".join(message_lines)
    
    results = []
    for num in numbers:
        try:
            message = twilio_client.messages.create(
                body=message_body,
                from_=f"whatsapp:{TWILIO_WHATSAPP_NUMBER}",
                to=f"whatsapp:{num['phone_number']}"
            )
            results.append({"phone": num["phone_number"], "status": "sent", "sid": message.sid})
            logger.info(f"Alert WhatsApp sent to {num['phone_number']}: {message.sid}")
        except Exception as e:
            results.append({"phone": num["phone_number"], "status": "failed", "error": str(e)})
            logger.error(f"Failed to send alert to {num['phone_number']}: {e}")
    
    # Log alert sent
    await db.alert_logs.insert_one({
        "id": str(uuid.uuid4()),
        "type": "whatsapp_ingredient_alert",
        "alerts_count": len(alerts),
        "recipients": len(numbers),
        "results": results,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    return {
        "message": f"Alertas enviadas a {len(numbers)} números",
        "alerts_count": len(alerts),
        "results": results
    }

async def get_critical_ingredient_alerts():
    """Get ingredients with critical stock (less than 3 days)"""
    inventory = await db.ingredient_inventory.find({}, {"_id": 0}).to_list(1000)
    ingredients = {i["id"]: i for i in await db.ingredients.find({}, {"_id": 0}).to_list(1000)}
    cafeterias = {c["id"]: c["name"] for c in await db.cafeterias.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(100)}
    
    # Calculate consumption from last 30 days
    thirty_days_ago = (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()
    movements = await db.ingredient_movements.find({
        "movement_type": "consumo_venta",
        "created_at": {"$gte": thirty_days_ago}
    }, {"_id": 0}).to_list(10000)
    
    daily_consumption = {}
    for mov in movements:
        key = f"{mov['ingredient_id']}_{mov.get('cafeteria_id', '')}"
        if key not in daily_consumption:
            daily_consumption[key] = 0
        daily_consumption[key] += mov["quantity"]
    
    critical_alerts = []
    for item in inventory:
        key = f"{item['ingredient_id']}_{item['cafeteria_id']}"
        avg_daily = daily_consumption.get(key, 0) / 30 if daily_consumption.get(key) else 0
        days_left = item["quantity"] / avg_daily if avg_daily > 0 else float('inf')
        
        # Critical = less than 3 days
        if days_left < 3 and days_left != float('inf'):
            ing = ingredients.get(item["ingredient_id"], {})
            critical_alerts.append({
                "ingredient_id": item["ingredient_id"],
                "ingredient_name": ing.get("name", "Desconocido"),
                "unit": ing.get("unit", "unidad"),
                "cafeteria_id": item["cafeteria_id"],
                "cafeteria_name": cafeterias.get(item["cafeteria_id"], "Desconocida"),
                "current_stock": item["quantity"],
                "min_stock": item["min_stock"],
                "avg_daily_consumption": round(avg_daily, 2),
                "days_until_stockout": round(days_left, 1),
                "suggested_order": round(avg_daily * 14 - item["quantity"], 2) if avg_daily > 0 else item["min_stock"] * 2
            })
    
    return sorted(critical_alerts, key=lambda x: x["days_until_stockout"])

# ============== SEED DATA ==============

@api_router.post("/seed")
async def seed_data():
    existing_cafes = await db.cafeterias.count_documents({})
    if existing_cafes > 0:
        return {"message": "Datos ya existentes"}
    
    # Create admin user
    admin_id = str(uuid.uuid4())
    admin = {
        "id": admin_id,
        "email": "admin@cafecontrol.com",
        "name": "Administrador",
        "password": hash_password("admin123"),
        "role": UserRole.ADMIN,
        "cafeteria_id": None,
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(admin)
    
    # Create cafeterias
    cafeterias_data = [
        {"name": "Doré Central", "address": "Av. Reforma 123, Centro", "phone": "555-0001"},
        {"name": "Doré Norte", "address": "Av. Insurgentes Norte 456", "phone": "555-0002"},
        {"name": "Doré Sur", "address": "Av. Universidad 789", "phone": "555-0003"}
    ]
    
    cafe_ids = []
    for cafe in cafeterias_data:
        cafe_id = str(uuid.uuid4())
        cafe_ids.append(cafe_id)
        await db.cafeterias.insert_one({
            **cafe,
            "id": cafe_id,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    
    # Create categories
    categories_data = [
        {"name": "Bebidas Calientes", "description": "Cafés, tés y bebidas calientes"},
        {"name": "Bebidas Frías", "description": "Frappés, smoothies y bebidas frías"},
        {"name": "Alimentos", "description": "Pasteles, sándwiches y snacks"},
        {"name": "Granos", "description": "Café en grano para venta"}
    ]
    
    cat_ids = []
    for cat in categories_data:
        cat_id = str(uuid.uuid4())
        cat_ids.append(cat_id)
        await db.categories.insert_one({"id": cat_id, **cat})
    
    # Create suppliers
    suppliers_data = [
        {"name": "Proveedora de Café MX", "contact_name": "Juan Pérez", "phone": "555-1001"},
        {"name": "Lácteos del Valle", "contact_name": "María López", "phone": "555-1002"},
        {"name": "Panadería Artesanal", "contact_name": "Carlos García", "phone": "555-1003"}
    ]
    
    supplier_ids = []
    for supplier in suppliers_data:
        supplier_id = str(uuid.uuid4())
        supplier_ids.append(supplier_id)
        await db.suppliers.insert_one({
            **supplier,
            "id": supplier_id,
            "email": None,
            "address": None,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    
    # Create ingredients
    ingredients_data = [
        {"name": "Café en grano", "unit": "kg", "cost_per_unit": 350.00, "supplier_id": supplier_ids[0], "min_stock": 5},
        {"name": "Leche entera", "unit": "litro", "cost_per_unit": 28.00, "supplier_id": supplier_ids[1], "min_stock": 20},
        {"name": "Leche deslactosada", "unit": "litro", "cost_per_unit": 32.00, "supplier_id": supplier_ids[1], "min_stock": 10},
        {"name": "Azúcar", "unit": "kg", "cost_per_unit": 25.00, "supplier_id": None, "min_stock": 5},
        {"name": "Jarabe de vainilla", "unit": "litro", "cost_per_unit": 180.00, "supplier_id": None, "min_stock": 3},
        {"name": "Jarabe de caramelo", "unit": "litro", "cost_per_unit": 180.00, "supplier_id": None, "min_stock": 3},
        {"name": "Chocolate en polvo", "unit": "kg", "cost_per_unit": 150.00, "supplier_id": None, "min_stock": 2},
        {"name": "Crema batida", "unit": "litro", "cost_per_unit": 85.00, "supplier_id": supplier_ids[1], "min_stock": 5},
        {"name": "Hielo", "unit": "kg", "cost_per_unit": 15.00, "supplier_id": None, "min_stock": 10},
    ]
    
    ingredient_ids = []
    for ing in ingredients_data:
        ing_id = str(uuid.uuid4())
        ingredient_ids.append(ing_id)
        await db.ingredients.insert_one({
            **ing,
            "id": ing_id,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    
    # Create products with costs
    products_data = [
        {"name": "Americano", "category_id": cat_ids[0], "price": 45.00, "cost": 8.00},
        {"name": "Latte", "category_id": cat_ids[0], "price": 55.00, "cost": 15.00},
        {"name": "Cappuccino", "category_id": cat_ids[0], "price": 55.00, "cost": 15.00},
        {"name": "Espresso", "category_id": cat_ids[0], "price": 35.00, "cost": 6.00},
        {"name": "Mocha", "category_id": cat_ids[0], "price": 65.00, "cost": 18.00},
        {"name": "Frappé Mocha", "category_id": cat_ids[1], "price": 75.00, "cost": 22.00},
        {"name": "Frappé Caramelo", "category_id": cat_ids[1], "price": 75.00, "cost": 22.00},
        {"name": "Smoothie Frutas", "category_id": cat_ids[1], "price": 65.00, "cost": 20.00},
        {"name": "Croissant", "category_id": cat_ids[2], "price": 45.00, "cost": 15.00},
        {"name": "Panini Jamón", "category_id": cat_ids[2], "price": 85.00, "cost": 35.00},
        {"name": "Cheesecake", "category_id": cat_ids[2], "price": 75.00, "cost": 25.00},
        {"name": "Café Grano 250g", "category_id": cat_ids[3], "price": 180.00, "cost": 80.00}
    ]
    
    prod_ids = []
    for prod in products_data:
        prod_id = str(uuid.uuid4())
        prod_ids.append(prod_id)
        await db.products.insert_one({
            **prod,
            "id": prod_id,
            "description": "",
            "is_active": True,
            "main_image": None,
            "images": [],
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    
    # Create recipes for some products
    recipes_data = [
        {
            "product_id": prod_ids[0],  # Americano
            "ingredients": [
                {"ingredient_id": ingredient_ids[0], "quantity": 0.018},  # 18g café
            ],
            "portions": 1,
            "auto_deduct": True
        },
        {
            "product_id": prod_ids[1],  # Latte
            "ingredients": [
                {"ingredient_id": ingredient_ids[0], "quantity": 0.018},  # 18g café
                {"ingredient_id": ingredient_ids[1], "quantity": 0.200},  # 200ml leche
            ],
            "portions": 1,
            "auto_deduct": True
        },
        {
            "product_id": prod_ids[4],  # Mocha
            "ingredients": [
                {"ingredient_id": ingredient_ids[0], "quantity": 0.018},  # 18g café
                {"ingredient_id": ingredient_ids[1], "quantity": 0.150},  # 150ml leche
                {"ingredient_id": ingredient_ids[6], "quantity": 0.020},  # 20g chocolate
                {"ingredient_id": ingredient_ids[7], "quantity": 0.030},  # 30ml crema
            ],
            "portions": 1,
            "auto_deduct": True
        },
        {
            "product_id": prod_ids[5],  # Frappé Mocha
            "ingredients": [
                {"ingredient_id": ingredient_ids[0], "quantity": 0.018},  # 18g café
                {"ingredient_id": ingredient_ids[1], "quantity": 0.150},  # 150ml leche
                {"ingredient_id": ingredient_ids[6], "quantity": 0.025},  # 25g chocolate
                {"ingredient_id": ingredient_ids[8], "quantity": 0.100},  # 100g hielo
                {"ingredient_id": ingredient_ids[7], "quantity": 0.040},  # 40ml crema
            ],
            "portions": 1,
            "auto_deduct": True
        },
    ]
    
    for recipe in recipes_data:
        await db.recipes.insert_one({
            **recipe,
            "id": str(uuid.uuid4()),
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    
    # Create inventory for each cafeteria (products and ingredients)
    for cafe_id in cafe_ids:
        # Product inventory
        for prod_id in prod_ids:
            await db.inventory.insert_one({
                "id": str(uuid.uuid4()),
                "product_id": prod_id,
                "cafeteria_id": cafe_id,
                "quantity": 50.0,
                "min_stock": 10.0,
                "unit": "unidad",
                "created_at": datetime.now(timezone.utc).isoformat()
            })
        
        # Ingredient inventory
        for i, ing_id in enumerate(ingredient_ids):
            initial_qty = [10, 50, 25, 10, 5, 5, 3, 10, 20][i]  # Different initial quantities
            await db.ingredient_inventory.insert_one({
                "id": str(uuid.uuid4()),
                "ingredient_id": ing_id,
                "cafeteria_id": cafe_id,
                "quantity": initial_qty,
                "min_stock": ingredients_data[i]["min_stock"],
                "created_at": datetime.now(timezone.utc).isoformat()
            })
    
    # Create sample sales
    import random
    for cafe_id in cafe_ids:
        for day_offset in range(7):
            num_sales = random.randint(5, 15)
            for _ in range(num_sales):
                sale_items = []
                num_items = random.randint(1, 4)
                selected_prods = random.sample(list(zip(prod_ids, products_data)), num_items)
                
                for pid, pdata in selected_prods:
                    qty = random.randint(1, 3)
                    sale_items.append({
                        "product_id": pid,
                        "product_name": pdata["name"],
                        "quantity": qty,
                        "unit_price": pdata["price"],
                        "subtotal": pdata["price"] * qty,
                        "cost": pdata["cost"] * qty
                    })
                
                subtotal = sum(i["subtotal"] for i in sale_items)
                cost_total = sum(i["cost"] for i in sale_items)
                tax = subtotal * 0.16
                
                sale_date = datetime.now(timezone.utc) - timedelta(days=day_offset, hours=random.randint(0, 12))
                
                await db.sales.insert_one({
                    "id": str(uuid.uuid4()),
                    "cafeteria_id": cafe_id,
                    "items": sale_items,
                    "subtotal": round(subtotal, 2),
                    "tax": round(tax, 2),
                    "total": round(subtotal + tax, 2),
                    "cost_total": round(cost_total, 2),
                    "profit": round(subtotal - cost_total, 2),
                    "payment_method": random.choice(["efectivo", "tarjeta", "clip"]),
                    "clip_transaction_id": None,
                    "notes": None,
                    "created_by": admin_id,
                    "created_at": sale_date.isoformat()
                })
    
    return {"message": "Datos de prueba creados exitosamente", "admin_credentials": {"email": "admin@cafecontrol.com", "password": "admin123"}}

# Root endpoint
@api_router.get("/")
async def root():
    return {"message": "Doré API v1.0"}

# Serve uploaded files
@api_router.get("/uploads/products/{filename}")
async def get_uploaded_image(filename: str):
    """Serve uploaded product images"""
    file_path = UPLOADS_DIR / "products" / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Imagen no encontrada")
    
    # Determine content type
    ext = filename.split(".")[-1].lower()
    content_types = {
        "jpg": "image/jpeg",
        "jpeg": "image/jpeg",
        "png": "image/png",
        "gif": "image/gif",
        "webp": "image/webp"
    }
    content_type = content_types.get(ext, "application/octet-stream")
    
    with open(file_path, "rb") as f:
        content = f.read()
    
    return Response(content=content, media_type=content_type)

# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
